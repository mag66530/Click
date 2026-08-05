"""
kp_audit.py – сверка КП с тем, что на самом деле заведено в Яндекс.Бизнесе.

Задача: взять лист КП, взять все организации аккаунта и ответить на три
вопроса по каждому городу.

  1. Есть ли в Яндексе организация для этого города?
  2. Не завелось ли их несколько (дубли – главная беда: посты и актуализация
     уходят в одну карточку, а вторая живёт своей жизнью)?
  3. Совпадают ли сайт, телефон и почта с тем, что записано в КП?

На выходе – то же самое КП, только с добавленными колонками. Исходные строки
не трогаем вообще: человек открывает файл и видит свою таблицу, просто шире.

Сеть в этом модуле не нужна: сюда приходят уже прочитанные строки КП и уже
собранный список организаций. Поэтому всю логику можно проверить тестами, что
и сделано в tests_audit.py.

Как сопоставляем город и организацию
────────────────────────────────────
Название у всех карточек одно («Авиапромсталь»), так что различает их только
адрес и сайт:

  • адрес Яндекса – «Южный федеральный округ, Ростовская область, городской
    округ Шахты, Шахты». Режем по запятым, у каждой части снимаем приставку
    («городской округ», «г.», «посёлок») и сравниваем с городом из КП;
  • сайт у проекта городской – shahty.aviastal.ru. Совпал хост – это тот же
    город, даже если адрес записан странно.

Совпадение по сайту весит больше, чем по адресу: сайт человек вписывает
руками в обе таблицы, а адрес Яндекс пишет сам и по-разному.
"""

from __future__ import annotations

import re
from typing import Any, Callable, Iterable

# Метка сборки: streamlit_app сверяет её и перезагружает модуль при расхождении.
# Метка сборки – одна на всё приложение, лежит в build.py. Держать её
# в каждом модуле своей строкой уже ломалось: у одного файла осталось
# старое значение, и Click принялся перезагружать все модули на каждое
# нажатие, пока не выбило по памяти.
from build import BUILD  # noqa: F401

# ─── Нормализация ───────────────────────────────────────────────────

# Приставки перед названием населённого пункта. Снимаем их, чтобы «городской
# округ Шахты» и «Шахты» стали одним и тем же.
_PLACE_PREFIX_RX = re.compile(
    r"^(?:"
    r"городской\s+округ|муниципальный\s+округ|городское\s+поселение|"
    r"сельское\s+поселение|внутригородской\s+район|"
    r"посёлок\s+городского\s+типа|поселок\s+городского\s+типа|пгт\.?|"
    r"рабочий\s+посёлок|рабочий\s+поселок|рп\.?|"
    r"город|г\.|гор\.|посёлок|поселок|пос\.|п\.|село|с\.|деревня|дер\.|д\.|"
    r"станица|ст-ца|ст\.|хутор|х\.|аул|микрорайон|мкр\.?"
    r")\s+",
    re.I,
)

# Части адреса, которые городом быть не могут. Без этого «улица Гагарина»
# рискует совпасть с городом Гагарин.
_NOT_A_PLACE_RX = re.compile(
    r"^(?:улица|ул\.|проспект|просп\.|пр-?кт|переулок|пер\.|шоссе|бульвар|б-р|"
    r"проезд|набережная|наб\.|площадь|пл\.|тупик|линия|аллея|тракт|квартал|"
    r"микрорайон|мкр|корпус|корп\.|строение|стр\.|дом|д\.|литера|этаж|офис|"
    r"помещение|здание|владение|км|территория)\b",
    re.I,
)

# Части адреса верхнего уровня: область, край, республика, округ. Городом не
# считаем, иначе «Ростовская область» перетянет на себя «Ростов-на-Дону».
_REGION_RX = re.compile(
    r"(?:\bобласть\b|\bкрай\b|\bреспублика\b|\bавтономн|федеральный\s+округ|"
    r"\bрайон\b|\bулус\b|\bобл\.|\bкр\.)",
    re.I,
)


def norm_text(s: str) -> str:
    """Строка для сравнения: без лишних пробелов, в нижнем регистре, ё → е."""
    s = re.sub(r"\s+", " ", str(s or "")).strip().lower()
    return s.replace("ё", "е")


def norm_city(name: str) -> str:
    """Название города для сравнения: без приставок, без ё, без хвостов в скобках."""
    s = norm_text(name)
    s = re.sub(r"\s*\([^)]*\)", "", s)          # «Шахты (склад)» → «Шахты»
    s = _PLACE_PREFIX_RX.sub("", s).strip()
    s = re.sub(r"[«»\"']", "", s)
    # Дефисы у Яндекса и в КП пишут разными знаками: Улан-Удэ, Улан‑Удэ.
    s = re.sub(r"[‑–—−]", "-", s)
    s = re.sub(r"\s*-\s*", "-", s)
    return s.strip(" .,")


def city_variants(name: str) -> list[str]:
    """
    Все написания города из одной ячейки КП.

    В таблице города со сменившимся именем записаны через косую черту:
    «Астана/Нур-Султан», «Сумгаит/Сумгайыт», «Нахичевань /Нахчыван». Яндекс
    знает одно из них – значит, проверять надо оба.
    """
    parts = re.split(r"\s*[/|]\s*", str(name or ""))
    out, seen = [], set()
    for p in parts:
        v = norm_city(p)
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return out


# Страна из КП и код региона у Яндекса. Нужны из-за одинаковых названий:
# Армавир есть и в России, и в Армении – без страны карточка одного города
# садится на строку другого.
COUNTRY_CODES = {
    "россия": "RU", "рф": "RU", "российская федерация": "RU",
    "казахстан": "KZ", "рк": "KZ",
    "беларусь": "BY", "белоруссия": "BY", "рб": "BY",
    "азербайджан": "AZ", "армения": "AM", "грузия": "GE",
    "киргизия": "KG", "кыргызстан": "KG", "узбекистан": "UZ",
    "таджикистан": "TJ", "туркменистан": "TM", "молдова": "MD", "молдавия": "MD",
    "украина": "UA",
}


def country_code(name: str) -> str:
    return COUNTRY_CODES.get(norm_text(name), "")


def address_places(address: str) -> set[str]:
    """
    Населённые пункты из адреса Яндекса.

    «Южный федеральный округ, Ростовская область, городской округ Шахты, Шахты»
    → {«шахты»}. Регионы и улицы отбрасываем.
    """
    out: set[str] = set()
    for raw in str(address or "").split(","):
        part = raw.strip()
        if not part or _NOT_A_PLACE_RX.match(part) or _REGION_RX.search(part):
            continue
        if re.fullmatch(r"[\d\W]+", part):        # «123», «116А» – номер дома
            continue
        place = norm_city(part)
        if place and not re.fullmatch(r"\d+[а-я]?", place):
            out.add(place)
    return out


def site_host(url: str) -> str:
    """Хост сайта без схемы, www и пути: https://Shahty.Aviastal.RU/ → shahty.aviastal.ru."""
    s = norm_text(url)
    if not s:
        return ""
    s = re.sub(r"^[a-z]+://", "", s)
    s = s.split("/")[0].split("?")[0].split("#")[0]
    s = re.sub(r"^www\.", "", s)
    return s.strip(" .")


def phone_key(phone: str) -> str:
    """
    Телефон для сравнения – последние 10 цифр.

    В КП пишут «7 (391) 271-75-19», Яндекс отдаёт «+7 (391) 271-75-19», а
    иногда номер записан вовсе без кода страны. Десять цифр хвоста совпадают
    во всех трёх случаях.
    """
    digits = re.sub(r"\D", "", str(phone or ""))
    if len(digits) < 6:
        return ""
    return digits[-10:] if len(digits) >= 10 else digits


def split_values(cell: str) -> list[str]:
    """Ячейка КП может держать несколько значений через запятую или перенос."""
    parts = re.split(r"[\n;,]+", str(cell or ""))
    return [p.strip() for p in parts if p.strip()]


def email_key(value: str) -> str:
    return norm_text(value).strip(" .,;")


SPRAV_ID_RX = re.compile(r"(?:sprav|/business/companies/company)/(\d+)", re.I)


def company_id_from_url(url: str) -> str:
    m = SPRAV_ID_RX.search(str(url or ""))
    return m.group(1) if m else ""


def company_ids(company: dict) -> set[str]:
    """
    Все номера, под которыми Яндекс знает одну и ту же карточку.

    В КП ссылки записаны старым номером (yandex.ru/sprav/25755702), а список
    организаций отдаёт новый (10873194809). Это одна карточка: старый номер
    лежит в tycoon_id. Пока сверка сравнивала только новый, почти на каждой
    строке всплывало «ссылка ведёт на другую карточку» – неправда.
    """
    return {str(company.get(k) or "") for k in ("id", "permanentId", "tycoonId")} - {""}


# ─── Разбор листа КП ────────────────────────────────────────────────

# Колонки ищем по названию. Названия в КП разных проектов немного разные,
# поэтому для каждой – несколько вариантов, сравнение по нормализованному
# тексту заголовка.
_COLS: dict[str, tuple[str, ...]] = {
    "country": ("страна",),
    "city": ("город", "населенный пункт", "нас. пункт"),
    "site": ("url", "сайт", "site", "адрес сайта"),
    "address": ("адрес", "адрес офиса"),
    "email": ("почта", "email", "e-mail", "эл. почта", "электронная почта"),
    "phone": ("общий город", "телефон", "тел.", "городской"),
    "phone2": ("общий сотовый", "сотовый", "мобильный"),
    "link": ("аккаунт", "ссылка", "яндекс", "яндекс бизнес"),
    "status": ("статус", "яндекс_статус"),
}

SPRAV_RX = re.compile(r"yandex\.[a-z.]+/sprav/\d+", re.I)

# Чужие площадки: их ссылки в КП тоже есть, но сайтом города они не являются.
_FOREIGN_SITE_RX = re.compile(r"yandex\.|2gis\.|2гис|google\.|goo\.gl|vk\.com|vk\.ru|"
                              r"ok\.ru|t\.me|wa\.me|instagram|facebook", re.I)


def _looks_like_site(value: str) -> bool:
    """Похоже ли значение ячейки на сайт города (а не на почту или чужую площадку)."""
    v = norm_text(value)
    if not v or "@" in v or " " in v.strip():
        return False
    if not re.search(r"[a-z0-9-]+\.[a-z]{2,}", v):
        return False
    return not _FOREIGN_SITE_RX.search(v)


def parse_sheet(rows: list[list[str]]) -> dict:
    """
    Разобрать лист КП: где шапка, какие колонки, какие строки – города.

    Возвращает {headerIdx, header, columns, items, width, error}. items –
    список {rowIdx, city, country, site, email, phones, link, status}, где
    rowIdx – номер строки в исходной таблице (по нему потом дописываем
    колонки, ничего не сдвигая).
    """
    out: dict[str, Any] = {
        "headerIdx": -1, "header": [], "columns": {}, "items": [],
        "width": max((len(r) for r in rows), default=0), "error": "",
    }
    if not rows:
        out["error"] = "Лист пустой"
        return out

    # Шапка – строка с «Город». «Страна» бывает не всегда (лист на одну страну).
    header_idx = -1
    for i, row in enumerate(rows[:25]):
        cells = [norm_text(c) for c in row]
        if any(c == "город" for c in cells):
            header_idx = i
            break
    if header_idx < 0:
        out["error"] = "Не нашли шапку: нужна строка с колонкой «Город»"
        return out

    header = [norm_text(c) for c in rows[header_idx]]
    out["headerIdx"] = header_idx
    out["header"] = rows[header_idx]

    def find(names: Iterable[str], start: int = 0) -> int:
        for want in names:
            for c in range(start, len(header)):
                if header[c] == want:
                    return c
        for want in names:                       # мягко: заголовок с уточнением
            for c in range(start, len(header)):
                if header[c].startswith(want):
                    return c
        return -1

    cols = {key: find(names) for key, names in _COLS.items()}
    data = rows[header_idx + 1:]

    # Городской сайт – по содержимому: в КП АПС эта колонка вообще без
    # заголовка, а сайт это главная зацепка при сопоставлении с Яндексом.
    if cols["site"] < 0:
        best, hits = -1, 0
        for c in range(out["width"]):
            n = sum(1 for r in data if c < len(r) and _looks_like_site(str(r[c])))
            if n > hits:
                best, hits = c, n
        if hits >= 3:
            cols["site"] = best

    # Ссылка на карточку: в КП колонок «Аккаунт» несколько (Яндекс, 2ГИС,
    # Google) – берём ту, где реально лежат ссылки на sprav.
    best_col, best_hits = -1, 0
    for c in range(out["width"]):
        hits = sum(1 for r in data if c < len(r) and SPRAV_RX.search(str(r[c])))
        if hits > best_hits:
            best_col, best_hits = c, hits
    if best_col >= 0:
        cols["link"] = best_col
        # «Статус» – ближайшая колонка справа от ссылок, но не чужой площадки.
        for c in range(best_col + 1, min(best_col + 6, len(header))):
            if "статус" in header[c] and not re.search(r"2гис|2gis|гугл|google", header[c]):
                cols["status"] = c
                break
    out["columns"] = cols

    def cell(row: list[str], idx: int) -> str:
        return str(row[idx]).strip() if 0 <= idx < len(row) else ""

    for n, row in enumerate(data):
        city = cell(row, cols["city"])
        if not city or norm_text(city) == "город":
            continue
        phones = split_values(cell(row, cols["phone"])) + split_values(cell(row, cols["phone2"]))
        out["items"].append({
            "rowIdx": header_idx + 1 + n,
            "city": city,
            "country": cell(row, cols["country"]),
            "site": cell(row, cols["site"]),
            "address": cell(row, cols["address"]),
            "email": cell(row, cols["email"]),
            "phones": phones,
            "link": cell(row, cols["link"]),
            "status": cell(row, cols["status"]),
        })
    return out


# ─── Сопоставление ──────────────────────────────────────────────────

SCORE_LINK = 100        # в КП стоит прямая ссылка на эту карточку – спорить не о чем
SCORE_SITE = 10         # совпал хост городского сайта
SCORE_PLACE = 6         # город назван в адресе отдельной частью
SCORE_COUNTRY = 2       # страна КП совпала с регионом карточки
SCORE_INSIDE = 2        # город лишь встречается в строке адреса


def match_score(item: dict, company: dict) -> int:
    """Насколько организация похожа на город КП. 0 – не похожа вовсе."""
    kp_id = company_id_from_url(item.get("link", ""))
    if kp_id and kp_id in company_ids(company):
        return SCORE_LINK

    # Разные страны – разговор окончен. Иначе армянский Армавир садится на
    # строку российского: название одно, и по нему они неотличимы.
    kp_country = country_code(item.get("country", ""))
    ya_country = (company.get("regionCode") or "").upper()
    if kp_country and ya_country and kp_country != ya_country:
        return 0

    score = SCORE_COUNTRY if (kp_country and kp_country == ya_country) else 0
    kp_host = site_host(item.get("site", ""))
    ya_hosts = {site_host(u) for u in (company.get("sites") or [company.get("site", "")])}
    ya_hosts.discard("")
    if kp_host and kp_host in ya_hosts:
        score += SCORE_SITE

    places = address_places(company.get("address", ""))
    address = norm_text(company.get("address", ""))
    for city in city_variants(item.get("city", "")):
        if city in places:
            score += SCORE_PLACE
            break
        # Совсем короткие названия («Ош», «РФ») по кусочку строки не ищем:
        # такое совпадение чаще случайное, чем настоящее.
        if len(city) >= 4 and re.search(rf"(?<![а-я]){re.escape(city)}(?![а-я])", address):
            score += SCORE_INSIDE
            break
    return score


def match(items: list[dict], companies: list[dict]) -> dict:
    """
    Разложить организации по городам КП.

    Каждая организация достаётся ОДНОМУ городу – тому, к которому подходит
    лучше всех. Так «Ростов-на-Дону» не утаскивает карточку «Ростова
    Великого»: у второго выше очки за точное совпадение части адреса.

    Возвращает {byRow: {rowIdx: [организации]}, extra: [...], chains: [...]}.
    """
    chains = [c for c in companies if (c.get("type") or "") == "chain"]
    ordinary = [c for c in companies if c not in chains]

    def longest(it: dict) -> int:
        return max((len(v) for v in city_variants(it.get("city", ""))), default=0)

    by_row: dict[int, list[dict]] = {it["rowIdx"]: [] for it in items}
    extra: list[dict] = []
    for co in ordinary:
        best_item, best_score = None, 0
        for it in items:
            s = match_score(it, co)
            if s > best_score or (s == best_score and s and best_item is not None
                                  and longest(it) > longest(best_item)):
                best_item, best_score = it, s
        # Одной страны мало: без совпадения по городу или сайту карточка
        # села бы на первую попавшуюся строку той же страны.
        if best_item is None or best_score <= SCORE_COUNTRY:
            extra.append(co)
            continue
        entry = dict(co)
        entry["matchScore"] = best_score
        entry["matchedBy"] = ("ссылка из КП" if best_score >= SCORE_LINK else
                              "сайт и адрес" if best_score >= SCORE_SITE + SCORE_PLACE else
                              "сайт" if best_score >= SCORE_SITE else
                              "адрес" if best_score >= SCORE_PLACE else "адрес (нестрого)")
        entry["matchedCountry"] = co.get("regionCode") or ""
        by_row[best_item["rowIdx"]].append(entry)

    for row in by_row.values():
        row.sort(key=lambda c: (-int(c.get("matchScore") or 0), str(c.get("id"))))
    return {"byRow": by_row, "extra": extra, "chains": chains}


# ─── Сравнение полей ────────────────────────────────────────────────

def _verdict(kp_values: list[str], ya_values: list[str], key: Callable[[str], str]) -> str:
    """
    Одно слово про поле: совпадает / расходится / нет в Яндексе / нет в КП.

    Совпадением считаем пересечение множеств: у Яндекса телефонов бывает
    несколько, и достаточно, чтобы номер из КП был среди них.
    """
    kp = {key(v) for v in kp_values if key(v)}
    ya = {key(v) for v in ya_values if key(v)}
    if not kp and not ya:
        return "–"
    if not ya:
        return "нет в Яндексе"
    if not kp:
        return "нет в КП"
    if kp & ya:
        return "совпадает" if kp <= ya else "совпадает частично"
    return "расходится"


BAD_VERDICTS = ("расходится", "нет в Яндексе")


def compare(item: dict, companies: list[dict]) -> dict:
    """Сравнить строку КП с найденными организациями. Ничего не меняет."""
    sites, emails, phones = [], [], []
    for co in companies:
        sites += [u for u in (co.get("sites") or []) if u]
        emails += [e for e in (co.get("emails") or []) if e]
        phones += [p for p in (co.get("phones") or []) if p]

    res = {
        "count": len(companies),
        "site": _verdict([item.get("site", "")], sites, site_host),
        "email": _verdict(split_values(item.get("email", "")), emails, email_key),
        "phone": _verdict(item.get("phones") or [], phones, phone_key),
        "yaSites": sites, "yaEmails": emails, "yaPhones": phones,
    }

    # Разногласия и подсказки держим врозь. Пустая ссылка в КП – это не
    # ошибка, а ровно то, что сверка и заполняет: красить из-за неё всю
    # таблицу в «что-то не так» бессмысленно, там таких строк почти все.
    problems: list[str] = []
    hints: list[str] = []
    if not companies:
        res["status"] = "нет"
        problems.append("в Яндексе организации не нашлось")
    elif len(companies) > 1:
        res["status"] = "несколько"
        problems.append(f"в Яндексе {len(companies)} карточки – проверьте дубли")
    else:
        res["status"] = "найдена"

    kp_id = company_id_from_url(item.get("link", ""))
    ya_ids: set[str] = set()
    for c in companies:
        ya_ids |= company_ids(c)
    # Старый сбор организаций не знал про tycoon_id – там сравнивать нечем,
    # и лучше промолчать, чем пугать ложной тревогой на каждой строке.
    knows_old_ids = all(("tycoonId" in c) for c in companies)
    if kp_id and ya_ids and knows_old_ids and kp_id not in ya_ids:
        problems.append(f"ссылка в КП ведёт на другую карточку ({kp_id})")
    if companies and not kp_id:
        hints.append("нет ссылки в КП – возьмите из колонки «Ссылка на карточку»")

    # Когда карточки нет вовсе, сравнивать не с чем: писать «сайт: нет в
    # Яндексе, телефон: нет в Яндексе, почта: нет в Яндексе» – это три раза
    # повторить то же самое и залить строку красным на ровном месте.
    if companies:
        for field, label in (("site", "сайт"), ("phone", "телефон"), ("email", "почта")):
            if res[field] in BAD_VERDICTS:
                problems.append(f"{label}: {res[field]}")
    for co in companies:
        if co.get("noAccess"):
            problems.append("карточка без доступа")
        if co.get("publishing") and co["publishing"] != "publish":
            problems.append(f"карточка не опубликована ({co['publishing']})")

    res["problems"] = problems
    res["hints"] = hints
    res["notes"] = problems + hints
    res["ok"] = res["status"] == "найдена" and not problems
    return res


# ─── Сборка отчёта ──────────────────────────────────────────────────

STATUS_MARK = {"найдена": "✅ найдена", "несколько": "❗ несколько", "нет": "❌ нет"}

# Колонки, которые дописываем справа к исходному КП. Порядок такой, чтобы
# слева стояло главное: что с карточкой, где она и что проверить. Раньше
# первыми шли справочные данные Яндекса, и до вывода приходилось листать.
EXTRA_HEADERS = [
    "Проверка", "Карточек", "Ссылка на карточку", "Что проверить",
    "Сайт совпадает?", "Телефон совпадает?", "Почта совпадает?",
    "Название в Яндексе", "Адрес в Яндексе", "Сайт в Яндексе",
    "Телефоны в Яндексе", "Почта в Яндексе", "Соцсети в Яндексе",
    "Рубрики в Яндексе",
]


def card_url(company_id: str) -> str:
    return f"https://yandex.ru/sprav/{company_id}/p/edit/" if company_id else ""


def _join(values: Iterable[str]) -> str:
    seen, out = set(), []
    for v in values:
        v = str(v or "").strip()
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return "\n".join(out)


def build(rows: list[list[str]], companies: list[dict]) -> dict:
    """
    Сверка целиком: разобрать КП, разложить организации, сравнить поля.

    Возвращает всё, что нужно и экрану, и выгрузке:
      sheet   – разбор листа,
      items   – строки КП с результатом сверки,
      extra   – организации, которым города в КП не нашлось,
      chains  – сетевые карточки (у них нет города, это «зонтик»),
      totals  – сводка цифрами.
    """
    sheet = parse_sheet(rows)
    if sheet["error"]:
        return {"sheet": sheet, "items": [], "extra": [], "chains": [],
                "totals": {}, "error": sheet["error"]}

    found = match(sheet["items"], companies)
    items = []
    totals = {"rows": len(sheet["items"]), "found": 0, "several": 0, "missing": 0,
              "mismatch": 0, "clean": 0, "noLink": 0, "extra": len(found["extra"]),
              "chains": len(found["chains"]), "companies": len(companies)}

    for it in sheet["items"]:
        cos = found["byRow"].get(it["rowIdx"], [])
        cmp = compare(it, cos)
        items.append({**it, "companies": cos, "cmp": cmp})
        if cmp["status"] == "нет":
            totals["missing"] += 1
        elif cmp["status"] == "несколько":
            totals["several"] += 1
            totals["found"] += 1
        else:
            totals["found"] += 1
        if cmp["ok"]:
            totals["clean"] += 1
        elif cmp["status"] != "нет":
            totals["mismatch"] += 1
        if cmp["hints"]:
            totals["noLink"] += 1

    return {"sheet": sheet, "items": items, "extra": found["extra"],
            "chains": found["chains"], "totals": totals, "error": ""}


def extra_cells(item: dict) -> list[str]:
    """Значения дописываемых колонок для одной строки КП."""
    cos = item.get("companies") or []
    cmp = item.get("cmp") or {}
    return [
        STATUS_MARK.get(cmp.get("status", ""), cmp.get("status", "")),
        str(len(cos)),
        _join(card_url(str(c.get("id") or "")) for c in cos),
        "; ".join(cmp.get("notes") or []),
        cmp.get("site", ""),
        cmp.get("phone", ""),
        cmp.get("email", ""),
        _join(c.get("name", "") for c in cos),
        _join(c.get("address", "") for c in cos),
        _join(cmp.get("yaSites") or []),
        _join(cmp.get("yaPhones") or []),
        _join(cmp.get("yaEmails") or []),
        _join(f"{k}: {v}" for c in cos for k, v in (c.get("social") or {}).items()),
        _join(r for c in cos for r in (c.get("rubrics") or [])),
    ]


def to_rows(rows: list[list[str]], result: dict) -> list[list[str]]:
    """
    Исходная таблица плюс новые колонки – ровно то, что просили: «то же КП,
    только со всей информацией из организаций».
    """
    sheet = result.get("sheet") or {}
    width = max(sheet.get("width", 0), max((len(r) for r in rows), default=0))
    by_row = {it["rowIdx"]: it for it in result.get("items") or []}
    header_idx = sheet.get("headerIdx", -1)

    out: list[list[str]] = []
    for i, row in enumerate(rows):
        line = [str(c) for c in row] + [""] * (width - len(row))
        if i == header_idx:
            line += EXTRA_HEADERS
        elif i in by_row:
            line += extra_cells(by_row[i])
        else:
            line += [""] * len(EXTRA_HEADERS)
        out.append(line)
    return out




# ─── Списки под конкретный вопрос ───────────────────────────────────
# «Ничего не понятно в отчёте» – потому что одна широкая простыня отвечала
# сразу на все вопросы. Теперь у каждого вопроса свой короткий список.

def guess_city(address: str) -> str:
    """
    Город из адреса Яндекса – для организаций, которых в КП нет.

    «Центральный федеральный округ, Ярославская область, городской округ
    Ярославль, Ярославль» → «Ярославль». Берём последнюю часть, которая
    похожа на населённый пункт: она же самая точная.
    """
    best = ""
    for raw in str(address or "").split(","):
        part = raw.strip()
        if not part or _NOT_A_PLACE_RX.match(part) or _REGION_RX.search(part):
            continue
        if re.fullmatch(r"[\d\W]+", part) or re.fullmatch(r"\d+[а-яa-z]?", norm_text(part)):
            continue
        best = _PLACE_PREFIX_RX.sub("", part).strip(" .,")
    return best


MISSING_HEADERS = ["Страна", "Город", "Сайт из КП", "Телефон из КП", "Почта из КП",
                   "Адрес из КП", "Статус в КП"]
EXTRA_HEADERS_SHEET = ["Город", "Название", "Адрес", "Сайт", "Телефоны", "Почта",
                       "Рубрики", "Публикация", "Ссылка на карточку"]
DOUBLE_HEADERS = ["Страна", "Город", "№", "Название", "Адрес", "Сайт", "Телефоны",
                  "Почта", "Ссылка на карточку"]
DIFF_HEADERS = ["Страна", "Город", "Что расходится", "В КП", "В Яндексе",
                "Ссылка на карточку"]


def missing_rows(result: dict) -> list[list[str]]:
    """Города КП, которых в Яндексе нет вовсе. Это список «что завести»."""
    out = [list(MISSING_HEADERS)]
    for it in result.get("items") or []:
        if (it.get("cmp") or {}).get("status") != "нет":
            continue
        out.append([it.get("country", ""), it.get("city", ""), it.get("site", ""),
                    " / ".join(it.get("phones") or []), it.get("email", ""),
                    it.get("address", ""), it.get("status", "")])
    return out


def extra_rows(result: dict) -> list[list[str]]:
    """
    Организации Яндекса, которых нет в КП, – со ссылкой на каждую.

    Заказчик просила именно ссылки: по ним видно, что это за карточка и что
    с ней делать – дописать город в КП или удалить дубль в Яндексе.
    """
    out = [list(EXTRA_HEADERS_SHEET)]
    for co in list(result.get("extra") or []) + list(result.get("chains") or []):
        out.append([
            "сеть (все города)" if co.get("type") == "chain" else guess_city(co.get("address", "")),
            co.get("name", ""),
            co.get("address", ""),
            _join(co.get("sites") or []),
            _join(co.get("phones") or []),
            _join(co.get("emails") or []),
            _join(co.get("rubrics") or []),
            "опубликована" if co.get("publishing") == "publish" else (co.get("publishing") or ""),
            card_url(str(co.get("id") or "")),
        ])
    return out


def double_rows(result: dict) -> list[list[str]]:
    """Города с несколькими карточками – по строке на карточку, чтобы выбрать лишнюю."""
    out = [list(DOUBLE_HEADERS)]
    for it in result.get("items") or []:
        cos = it.get("companies") or []
        if len(cos) < 2:
            continue
        for n, co in enumerate(cos, 1):
            out.append([it.get("country", ""), it.get("city", ""), str(n),
                        co.get("name", ""), co.get("address", ""),
                        _join(co.get("sites") or []), _join(co.get("phones") or []),
                        _join(co.get("emails") or []), card_url(str(co.get("id") or ""))])
    return out


def diff_rows(result: dict) -> list[list[str]]:
    """Расхождения по полям – по строке на поле, рядом оба значения."""
    out = [list(DIFF_HEADERS)]
    for it in result.get("items") or []:
        cmp = it.get("cmp") or {}
        if cmp.get("status") == "нет":
            continue
        link = card_url(str((it.get("companies") or [{}])[0].get("id") or ""))
        for field, label, kp_value, ya_value in (
            ("site", "Сайт", it.get("site", ""), " / ".join(cmp.get("yaSites") or [])),
            ("phone", "Телефон", " / ".join(it.get("phones") or []),
             " / ".join(cmp.get("yaPhones") or [])),
            ("email", "Почта", it.get("email", ""), " / ".join(cmp.get("yaEmails") or [])),
        ):
            if cmp.get(field) in BAD_VERDICTS:
                out.append([it.get("country", ""), it.get("city", ""),
                            f"{label}: {cmp[field]}", kp_value, ya_value, link])
    return out


def to_csv(rows: list[list[str]]) -> str:
    import csv
    import io

    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", lineterminator="\n")
    for row in rows:
        w.writerow([str(c).replace("\n", " / ") for c in row])
    return buf.getvalue()




# ─── Excel ──────────────────────────────────────────────────────────
# Оформление то же, что в отчёте site-checker «Проверка КП»: заказчик его уже
# читает, и второй документ в другой манере пришлось бы осваивать заново.
# Оттуда же взяты главные приёмы: в сетке стоят значки ✓ ✗ ⚠ –, а подробности
# висят примечанием к ячейке; заливка только у проблемных; листы «Дашборд» и
# «Как читать» идут первыми.

SYMBOL = {"ok": "✓", "bug": "✗", "warn": "⚠", "na": "–"}
COLOR = {"ok": "1E8E3E", "bug": "C62828", "warn": "B26A00", "na": "9E9E9E"}

C_HEAD = "EEF3FB"        # шапка
C_BUG = "FDE3E3"         # мягкий красный
C_WARN = "FFF2DA"        # мягкий оранжевый
C_OK = "E6F4EA"          # мягкий зелёный
C_LINK = "1155CC"
C_GREY = "595959"
C_BORDER = "C9CFDB"

# Как слово из сравнения превращается в значок.
VERDICT_KIND = {
    "совпадает": "ok",
    "совпадает частично": "warn",
    "нет в КП": "warn",
    "расходится": "bug",
    "нет в Яндексе": "bug",
    "–": "na",
}
STATUS_KIND = {"найдена": "ok", "несколько": "warn", "нет": "bug"}

LEGEND = [
    ("Как читать отчёт", True),
    ("", False),
    ("Сверка отвечает на один вопрос: то ли самое заведено в Яндекс.Бизнесе, "
     "что записано в КП. Организации читаются из раздела «Организации» вашего "
     "аккаунта – адрес, сайт, телефоны и почту Яндекс отдаёт там сразу.", False),
    ("", False),
    ("Значки на листе «Сверка»", True),
    ("✓  – в КП и в карточке одно и то же (для телефона: номер из КП есть среди "
     "номеров карточки).", False),
    ("✗  – расхождение: в карточке другое значение либо его там нет вовсе.", False),
    ("⚠  – сходится не полностью: в карточке есть лишнее значение, либо в КП "
     "пусто, а в Яндексе что-то указано.", False),
    ("–  – сверять нечего: ни в КП, ни в карточке ничего нет.", False),
    ("Наведите курсор на ячейку со значком – в примечании видно, что в КП и что "
     "в Яндексе.", False),
    ("", False),
    ("Колонка «Карточка»", True),
    ("✓ – карточка одна, всё в порядке. ⚠ – карточек несколько, это дубли: "
     "посты и актуализация уходят в одну, вторая живёт своей жизнью. "
     "✗ – карточки нет вовсе.", False),
    ("", False),
    ("Листы отчёта", True),
    ("Дашборд – сводка цифрами: сколько нашлось, где дубли, где расхождения.", False),
    ("Нет в КП – организации есть в Яндексе, а города в КП нет. Со ссылкой на "
     "каждую карточку.", False),
    ("Дубли – города, где карточек больше одной. По строке на карточку.", False),
    ("Нет в Яндексе – города из КП, для которых карточки не нашлось.", False),
    ("Расхождения – каждое расхождение отдельной строкой: слева КП, справа Яндекс.", False),
    ("Сверка – сетка ✓/✗ по всем городам, одна строка на город.", False),
    ("КП с данными – ваша таблица без единого изменения плюс колонки из Яндекса "
     "справа. Нужна, когда хочется видеть всё разом.", False),
]


def _bar(value: int, top: int, width: int = 18) -> str:
    """Полоска из блоков вместо графика: рисуется везде и без единой зависимости."""
    if top <= 0 or value <= 0:
        return ""
    filled = max(1, round(width * value / top))
    return "█" * filled


def _autosize(ws, first_row: int = 1, cap: int = 46) -> None:
    from openpyxl.utils import get_column_letter

    for col in range(1, ws.max_column + 1):
        longest = 0
        for row in range(first_row, min(ws.max_row, 500) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            longest = max(longest, max((len(p) for p in str(v).split("\n")), default=0))
        ws.column_dimensions[get_column_letter(col)].width = min(max(longest + 2, 9), cap)


def _style_head(ws, row: int) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color=C_BORDER)
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=C_HEAD)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def _linkify(ws, col: int, first_row: int, label: str = "") -> None:
    from openpyxl.styles import Font

    for row in range(first_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        url = str(cell.value or "").strip().split("\n")[0]
        if not url.startswith("http"):
            continue
        if label:
            cell.value = label
        cell.hyperlink = url
        cell.font = Font(color=C_LINK, underline="single")


def _list_sheet(wb, title: str, rows: list[list[str]], note: str = "",
                link_col: int | None = None, link_label: str = "",
                empty_note: str = "Пусто – и это хорошо.") -> None:
    """Лист-список: пояснение, шапка, фильтр, закреплённая шапка, ссылки."""
    from openpyxl.styles import Font

    ws = wb.create_sheet(title[:31])
    head_row = 1
    if note:
        ws.append([note])
        ws["A1"].font = Font(italic=True, color=C_GREY)
        ws.append([])
        head_row = 3
    for line in rows:
        ws.append(line)
    _style_head(ws, head_row)
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)
    if ws.max_row > head_row:
        last = ws.cell(row=head_row, column=ws.max_column).coordinate
        ws.auto_filter.ref = f"A{head_row}:{last}"
    _autosize(ws, first_row=head_row)
    if link_col:
        _linkify(ws, link_col, head_row + 1, link_label)
        if link_label:
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(link_col)].width = max(len(link_label) + 4, 18)
    if len(rows) <= 1:
        ws.cell(row=head_row + 1, column=1, value=empty_note)


def _dashboard(ws, result: dict, sheet_name: str, when: str) -> None:
    """Первый лист: цифры крупно, под ними – по какому полю сколько расхождений."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    t = result.get("totals") or {}
    thin = Side(style="thin", color=C_BORDER)
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    ws.sheet_view.showGridLines = False

    ws["B2"] = "Сверка КП с организациями Яндекса"
    ws["B2"].font = Font(size=16, bold=True, color=C_LINK)
    ws.merge_cells("B3:H3")
    ws["B3"] = (f"Лист КП «{sheet_name}»"
                + (f" · организации прочитаны {when}" if when else "")
                + f" · организаций в аккаунте: {t.get('companies', 0)}")
    ws["B3"].font = Font(size=10, color=C_GREY)

    tiles = (
        ("B5", "Городов в КП", "B6", t.get("rows", 0), C_LINK),
        ("D5", "Карточка нашлась", "D6", t.get("found", 0), "1E8E3E"),
        ("F5", "Нет в Яндексе", "F6", t.get("missing", 0), "C62828"),
        ("H5", "Есть в Яндексе, нет в КП", "H6", t.get("extra", 0), "9C7A00"),
        ("B8", "Всё сходится", "B9", t.get("clean", 0), "1E8E3E"),
        ("D8", "Несколько карточек", "D9", t.get("several", 0), "B26A00"),
        ("F8", "Есть расхождения", "F9", t.get("mismatch", 0), "C62828"),
        ("H8", "Без ссылки в КП", "H9", t.get("noLink", 0), C_GREY),
    )
    for lbl_at, label, val_at, value, colour in tiles:
        ws[lbl_at] = label
        ws[lbl_at].font = Font(size=9, color=C_GREY)
        ws[lbl_at].alignment = Alignment(wrap_text=True, vertical="bottom")
        ws[val_at] = value
        ws[val_at].font = Font(size=20, bold=True, color=colour)

    # Что именно расходится – с полоской, чтобы перекос был виден без графика.
    by_field = {"Сайт": 0, "Телефон": 0, "Почта": 0}
    for it in result.get("items") or []:
        cmp = it.get("cmp") or {}
        for key, label in (("site", "Сайт"), ("phone", "Телефон"), ("email", "Почта")):
            if cmp.get(key) in BAD_VERDICTS:
                by_field[label] += 1

    ws.merge_cells("B11:H11")
    ws["B11"] = "Расхождения по типу данных"
    ws["B11"].font = Font(size=12, bold=True)
    for c, title in enumerate(("Что сверяли", "Расхождений", ""), 2):
        cell = ws.cell(row=12, column=c, value=title)
        cell.font = Font(size=10, bold=True)
        cell.fill = PatternFill("solid", fgColor=C_HEAD)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    top = max(by_field.values(), default=0)
    for n, (label, value) in enumerate(by_field.items(), start=13):
        ws.cell(row=n, column=2, value=label).border = border
        ws.cell(row=n, column=3, value=value).border = border
        bar = ws.cell(row=n, column=4, value=_bar(value, top))
        bar.font = Font(color="C62828" if value else C_GREY)

    ws.append([])
    row = ws.max_row + 1
    ws.cell(row=row, column=2, value="Подробности – на листах «Нет в КП», «Дубли», "
                                     "«Нет в Яндексе» и «Расхождения». Как читать значки – "
                                     "на листе «Как читать».").font = Font(color=C_GREY)

    for col, width in (("A", 2), ("B", 24), ("C", 13), ("D", 22), ("E", 13),
                       ("F", 22), ("G", 13), ("H", 24)):
        ws.column_dimensions[col].width = width
    ws.row_dimensions[5].height = 26
    ws.row_dimensions[8].height = 26


def _legend_sheet(ws) -> None:
    from openpyxl.styles import Alignment, Font

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 104
    for i, (text, bold) in enumerate(LEGEND, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=14 if (bold and i == 1) else 11)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


GRID_HEADERS = ["Страна", "Город", "Карточка", "Сайт", "Телефон", "Почта",
                "Что проверить", "Ссылка на карточку", "Ошибок (✗)"]


def _grid_sheet(wb, result: dict) -> None:
    """
    Главный лист: одна строка на город, значки вместо простыней текста.

    Подробности – в примечании к ячейке, как в отчёте site-checker: там это
    единственное, что делает таблицу на две сотни строк читаемой.
    """
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    ws = wb.create_sheet("Сверка")
    for c, title in enumerate(GRID_HEADERS, 1):
        ws.cell(row=1, column=c, value=title)
    _style_head(ws, 1)
    ws.freeze_panes = "C2"
    thin = Side(style="thin", color=C_BORDER)
    border = Border(right=thin)

    def mark(row: int, col: int, kind: str, note: str = "") -> None:
        cell = ws.cell(row=row, column=col, value=SYMBOL[kind])
        cell.font = Font(color=COLOR[kind], bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        if kind in ("bug", "warn") and note:
            cell.fill = PatternFill("solid", fgColor=C_BUG if kind == "bug" else C_WARN)
        if note:
            comment = Comment(note[:400], "Click")
            comment.width, comment.height = 330, 150
            cell.comment = comment

    r = 1
    for it in result.get("items") or []:
        r += 1
        cmp = it.get("cmp") or {}
        cos = it.get("companies") or []
        ws.cell(row=r, column=1, value=it.get("country", ""))
        ws.cell(row=r, column=2, value=it.get("city", ""))

        status = cmp.get("status", "нет")
        card_note = {"найдена": "Карточка одна – всё в порядке.",
                     "несколько": f"Карточек {len(cos)} – это дубли, проверьте лишнюю.",
                     "нет": "Карточки в Яндексе не нашлось."}.get(status, "")
        if cos:
            card_note += "\n\n" + "\n".join(
                f'{c.get("name", "")} · {c.get("address", "")}'.strip(" ·") for c in cos)
        mark(r, 3, STATUS_KIND.get(status, "bug"), card_note)

        for n, (key, label, kp_value, ya_value) in enumerate((
            ("site", "Сайт", it.get("site", ""), " / ".join(cmp.get("yaSites") or [])),
            ("phone", "Телефон", " / ".join(it.get("phones") or []),
             " / ".join(cmp.get("yaPhones") or [])),
            ("email", "Почта", it.get("email", ""), " / ".join(cmp.get("yaEmails") or [])),
        )):
            # Карточки нет – сверять не с чем, ставим «–». Иначе строка
            # ненайденного города краснела четырьмя крестами про одно и то же.
            verdict = cmp.get(key, "–") if cos else "–"
            kind = VERDICT_KIND.get(verdict, "na")
            note = "" if kind == "na" else (f"{label}: {verdict}\n\n"
                                            f"КП: {kp_value or '–'}\n"
                                            f"Яндекс: {ya_value or '–'}")
            mark(r, 4 + n, kind, note)

        ws.cell(row=r, column=7, value="; ".join(cmp.get("problems") or []))
        ws.cell(row=r, column=8, value=card_url(str(cos[0].get("id"))) if cos else "")
        errors = len(cmp.get("problems") or [])
        err = ws.cell(row=r, column=9, value=errors)
        err.alignment = Alignment(horizontal="center")
        if errors:
            err.font = Font(bold=True, color="C62828")

    _linkify(ws, 8, 2, "Открыть карточку")
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(GRID_HEADERS)).coordinate}"
    for col, width in (("A", 14), ("B", 24), ("C", 11), ("D", 9), ("E", 11),
                       ("F", 9), ("G", 52), ("H", 20), ("I", 11)):
        ws.column_dimensions[col].width = width
    if r == 1:
        ws.cell(row=2, column=1, value="В листе КП не нашлось ни одного города.")


def _full_sheet(wb, rows: list[list[str]], result: dict) -> None:
    """Исходное КП плюс колонки из Яндекса – для тех, кому нужна вся таблица."""
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet("КП с данными")
    body = to_rows(rows, result)
    for line in body:
        ws.append(line)

    sheet = result.get("sheet") or {}
    header_idx = sheet.get("headerIdx", -1)
    status_col = sheet.get("width", 0) + 1
    city_col = (sheet.get("columns") or {}).get("city", 0) + 1
    fills = {"✅ найдена": C_OK, "❗ несколько": C_WARN, "❌ нет": C_BUG}

    if header_idx >= 0:
        _style_head(ws, header_idx + 1)
        ws.freeze_panes = ws.cell(row=header_idx + 2, column=max(city_col + 1, 2))
        last = ws.cell(row=header_idx + 1, column=ws.max_column).coordinate
        ws.auto_filter.ref = f"A{header_idx + 1}:{last}"

    # Красим только саму плашку статуса и разошедшиеся поля: заливка на все
    # четырнадцать колонок рябила так, что читать было нечем.
    diff_cols = [status_col + n for n, name in enumerate(EXTRA_HEADERS)
                 if name.endswith("совпадает?")]
    for i, line in enumerate(body, start=1):
        mark = line[status_col - 1] if status_col - 1 < len(line) else ""
        if mark in fills:
            cell = ws.cell(row=i, column=status_col)
            cell.fill = PatternFill("solid", fgColor=fills[mark])
            cell.font = Font(bold=True)
        for col in diff_cols:
            if col - 1 < len(line) and line[col - 1] in BAD_VERDICTS:
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=C_BUG)

    _linkify(ws, status_col + EXTRA_HEADERS.index("Ссылка на карточку"),
             (header_idx + 2) if header_idx >= 0 else 2, "Открыть карточку")
    _autosize(ws, first_row=max(header_idx + 1, 1), cap=38)
    for c in range(status_col, status_col + len(EXTRA_HEADERS)):
        ws.cell(row=max(header_idx + 1, 1), column=c).alignment = Alignment(
            wrap_text=True, vertical="center", horizontal="center")


def to_xlsx(rows: list[list[str]], result: dict, sheet_name: str = "Сверка",
            collected_at: str = "") -> bytes:
    """
    Отчёт книгой: дашборд, подсказка, три списка под конкретный вопрос,
    сетка ✓/✗ и полное КП.

    Раньше это была одна таблица в сорок колонок, и заказчик прямо сказала:
    «ничего не понятно». Один лист – один вопрос; широкая таблица лежит
    последней, для случаев, когда нужна именно она.
    """
    import io

    from openpyxl import Workbook

    wb = Workbook()
    _dashboard(wb.active, result, sheet_name, collected_at)
    wb.active.title = "Дашборд"
    _legend_sheet(wb.create_sheet("Как читать"))

    _list_sheet(wb, "Нет в КП", extra_rows(result), link_col=9, link_label="Открыть карточку",
                note="Эти организации заведены в Яндексе, а города в КП нет. "
                     "Ссылка ведёт прямо в карточку: дописать город в КП или удалить дубль.",
                empty_note="Все организации разошлись по городам КП.")
    _list_sheet(wb, "Дубли", double_rows(result), link_col=9, link_label="Открыть карточку",
                note="Города, где карточек больше одной. Строк столько же, сколько карточек.",
                empty_note="Дублей нет.")
    _list_sheet(wb, "Нет в Яндексе", missing_rows(result),
                note="Города из КП, для которых карточки в Яндексе не нашлось.",
                empty_note="Все города КП нашлись в Яндексе.")
    _list_sheet(wb, "Расхождения", diff_rows(result), link_col=6, link_label="Открыть карточку",
                note="Слева значение из КП, справа – из карточки Яндекса.",
                empty_note="Расхождений не найдено 🎉")

    _grid_sheet(wb, result)
    _full_sheet(wb, rows, result)
    wb.active = 0

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
