"""
content_plan.py – чтение реестра постов (контент-плана) из таблицы бренда.

Зачем. Кросспостинг исполняет реестр: строки с датой, текстом, картинкой и
списком соцсетей. Раньше это была памятка для человека – теперь исполняемый
план. Этот модуль превращает таблицу в список постов, из которых планировщик
делает отложки и задания.

Откуда данные. Та же Google-таблица и тот же сервисный аккаунт, что у КП
(`kp_sheet.py`): заказчик расшаривает файл реестра на него Читателем – новых
ключей не нужно. Для локальной проверки читаем и обычный .xlsx с диска.

Структура реестра (файл заказчика, разобран 2026-08-07). Лист на бренд
(СМУ/ИМП/МПЭ/АПС/…). Пост – БЛОК строк:
  • первая строка блока несёт ДАТУ, ТЕКСТ, ФОТО, ТИП;
  • каждая строка блока (включая первую) – одна СОЦСЕТЬ (колонка «Соцсеть»)
    со своей колонкой «Ссылка» на опубликованный пост;
  • у будущих постов «Ссылка» пуста – это и есть «ещё не сформировано».
Справа в листе отдельный блок статистики – его не трогаем (читаем только
колонки, найденные по заголовкам «Когда выложить», «Соцсеть», «Ссылка»,
«Формат», «Тип», «Пост», «Фото»).

Время. В реестре только дата. Час выхода – из настройки бренда (заводские
значения ниже), по Екатеринбургу – как всё видимое время в Click (`apptime`).
"""

from __future__ import annotations

import re
from datetime import date, datetime
from typing import Any, Callable

import apptime

# Метка сборки – одна на всё приложение (см. build.py). streamlit_app сверяет
# её и перезагружает модуль при расхождении.
from build import BUILD  # noqa: F401

# ─── Бренды ─────────────────────────────────────────────────────────
# Заводской час публикации на бренд (Екатеринбург). Ответ заказчика 2026-08-07.
# Правится в «Настройках»; переопределяется колонкой времени в реестре, если
# заказчик её заведёт.
DEFAULT_TIMES = {
    "SMU": "11:00",
    "IMP": "09:00",
    "MPE": "10:00",
    "MPI": "09:30",
    "APS": "10:30",
}

# Название листа в реестре → код проекта Click. Лист МПИ заказчик добавит –
# парсер к нему готов, отдельного кода не нужно.
SHEET_TO_BRAND = {
    "СМУ": "SMU",
    "ИМП": "IMP",
    "МПЭ": "MPE",
    "МПИ": "MPI",
    "АПС": "APS",
}

# ─── Соцсети ────────────────────────────────────────────────────────
# Как площадка называется в реестре → канонический код цели. У Телеграма два
# канала на бренд, поэтому tg-staff и tg-client различаем.
#
# SUPPORTED – то, что Click формирует САМ. Телеграм теперь здесь: два его
# канала (tg-staff, tg-client) Click ставит РОДНОЙ отложкой через веб-версию
# под аккаунтом-админом – так же, как ВК/ОК/МАКС/Дзен. Раньше их тут не было
# (13.08.2026): у бота отложки нет, время держал сам Click, пост уходил не в
# том виде. С браузерной отложкой это позади – время держит сам Телеграм.
#
# Обобщённый «tg» (реестр написал просто «Telegram», не различив клиентов и
# сотрудников) в SUPPORTED НЕ вносим: непонятно, в какой из двух каналов
# ставить. Такая цель остаётся «вручную», пока в реестре её не уточнят.
TELEGRAM = ("tg-staff", "tg-client", "tg")
SUPPORTED = ("vk", "ok", "max", "zen", "tg-staff", "tg-client")

# Площадки, которые берут СТАТЬИ, а не посты. Дзен – единственная такая:
# туда уходит лонгрид, а текст его лежит не в ячейке реестра, а по ссылке на
# документ (см. zen_doc). Поэтому строка формата «Статья» – это работа для
# Дзена и НЕ работа для ВК с МАКСом: им статья не нужна, у них свой короткий
# пост в той же таблице.
ARTICLE_NETWORKS = ("zen",)


def canonical_network(raw: str) -> str:
    """Название соцсети из реестра → код цели. Пусто, если площадка не наша."""
    s = (raw or "").strip().lower()
    if not s:
        return ""
    if "вконтакт" in s or s == "вк":
        return "vk"
    if "одноклас" in s or s == "ок":            # в реестре встречается «Однокласники»
        return "ok"
    if "telegram" in s or "телеграм" in s or s == "тг":
        if "сотруд" in s:
            return "tg-staff"
        if "клиент" in s:
            return "tg-client"
        return "tg"
    if "max" in s or "макс" in s:
        return "max"
    if "дзен" in s or "zen" in s:
        return "zen"
    return ""


# ─── Заголовки колонок ──────────────────────────────────────────────
# Колонку ищем по заголовку, а не по букве: так порядок колонок можно менять.
_COL_MATCHERS: dict[str, Callable[[str], bool]] = {
    "date":  lambda s: "когда выложить" in s or s == "дата",
    "net":   lambda s: "соцсет" in s or "соц.сет" in s,
    "link":  lambda s: s == "ссылка",
    "format": lambda s: "формат" in s,
    "type":  lambda s: s == "тип",
    "text":  lambda s: s == "пост" or s == "текст",
    "photo": lambda s: "фото" in s or "картинк" in s,
}


def _find_header(rows: list[list[str]]) -> tuple[int, dict[str, int]]:
    """Строка заголовков и карта «поле → индекс колонки». (-1, {}) – не нашли."""
    for i, row in enumerate(rows[:30]):        # шапка всегда наверху
        # Шапки в таблицах сплошь жирные – маркеры разметки тут только мешают.
        norm = [(c or "").replace("**", "").strip().lower() for c in row]
        has_date = any(_COL_MATCHERS["date"](c) for c in norm)
        has_net = any(_COL_MATCHERS["net"](c) for c in norm)
        if not (has_date and has_net):
            continue
        cols: dict[str, int] = {}
        for field, match in _COL_MATCHERS.items():
            for j, c in enumerate(norm):
                if match(c):
                    cols[field] = j
                    break
        return i, cols
    return -1, {}


# ─── Дата и время ───────────────────────────────────────────────────
_DATE_FORMATS = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d",
                 "%d.%m.%Y", "%d.%m.%y", "%d/%m/%Y")


def parse_date(value: str) -> date | None:
    """Дата из ячейки. Разные источники пишут её по-разному – пробуем набор форматов."""
    s = (value or "").strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def brand_default_time(brand: str) -> str:
    return DEFAULT_TIMES.get(brand, "10:00")


def when_iso(d: date, hhmm: str) -> str:
    """Дата + час бренда → ISO с поясом Екатеринбурга («2026-08-12T11:00:00+05:00»)."""
    hh, mm = (hhmm or "10:00").split(":")[:2]
    dt = datetime(d.year, d.month, d.day, int(hh), int(mm), tzinfo=apptime.TZ)
    return dt.isoformat()


# ─── Картинки ───────────────────────────────────────────────────────
def _split_images(cell: str) -> list[str]:
    """
    Ссылки на фото из ячейки (разделители: перенос, пробел, запятая, ;).

    Заказчица держит в ячейке ДВА варианта, помеченные словами: «ЛОГО: <url>»
    и «БЕЗ ЛОГО: <url>» (картинка с логотипом на ней или без). Постим ВСЕГДА
    вариант С логотипом: если в ячейке есть пометка «ЛОГО», берём адреса из
    строк с ней и выкидываем строки «БЕЗ ЛОГО». Нет пометок – как раньше: все
    ссылки подряд. Слово «ЛОГО» ищем по-русски, поэтому latin «logo» в самом
    адресе (имя файла) отбор не путает.
    """
    text = (cell or "").strip()
    if not text:
        return []
    lines = [ln for ln in text.splitlines() if ln.strip()]
    if any(re.search(r"лого", ln, re.I) for ln in lines):
        lines = [ln for ln in lines
                 if re.search(r"лого", ln, re.I)
                 and not re.search(r"без\s*лого", ln, re.I)]
    parts = re.split(r"[\s,;]+", "\n".join(lines))
    return [p for p in parts if p.startswith("http")]


# ─── Разбор листа ───────────────────────────────────────────────────
def parse_sheet(rows: list[list[str]], brand: str) -> list[dict]:
    """
    Лист реестра → список постов. Каждый пост:
      {brand, date (ISO), time (HH:MM), when (ISO+пояс), post_type, format,
       text, images:[...], targets:[{network, raw, published_link}], row}
    Возвращаются ВСЕ посты (и прошлые, и будущие); что именно формировать –
    решает posts_to_form(). Так проще и тестировать, и показывать план целиком.
    """
    hdr, cols = _find_header(rows)
    if hdr < 0 or "date" not in cols or "net" not in cols:
        return []

    def cell(row: list[str], field: str) -> str:
        j = cols.get(field, -1)
        raw = row[j].strip() if 0 <= j < len(row) else ""
        # Разметка жирного (**…**) имеет смысл только в тексте поста. В прочих
        # колонках это шум: жирная дата или жирное «Вконтакте» в реестре не
        # должны ломать распознавание. Снимаем маркеры везде, кроме текста.
        return raw if field == "text" else raw.replace("**", "").strip()

    time = brand_default_time(brand)
    posts: list[dict] = []
    cur: dict | None = None

    for idx in range(hdr + 1, len(rows)):
        row = rows[idx]
        d = parse_date(cell(row, "date"))
        net_raw = cell(row, "net")

        if d is not None:
            # новая строка с датой – начало нового блока-поста
            cur = {
                "brand": brand,
                "date": d.isoformat(),
                "time": time,
                "when": when_iso(d, time),
                "post_type": cell(row, "type"),
                "format": cell(row, "format") or "Пост",
                "text": cell(row, "text"),
                "images": _split_images(cell(row, "photo")),
                "targets": [],
                "row": idx + 1,               # 1-based, как в редакторе таблиц
            }
            posts.append(cur)

        if net_raw and cur is not None:
            cur["targets"].append({
                "network": canonical_network(net_raw),
                "raw": net_raw,
                "published_link": cell(row, "link"),
            })

    return [p for p in posts if p["targets"]]


# ─── Что формировать ────────────────────────────────────────────────
def is_article(post: dict) -> bool:
    """Строка реестра – статья (лонгрид), а не пост в ленту."""
    return (post.get("format") or "Пост").strip().lower() != "пост"


def posts_to_form(posts: list[dict], today: date | None = None) -> list[dict]:
    """
    Отобрать посты и цели, которые ещё надо сформировать:
      • дата сегодня или впереди;
      • есть текст (у статьи – ссылка на документ, это тоже текст ячейки);
      • цель – наша площадка (SUPPORTED) и «Ссылка» по ней пуста;
      • формат и площадка друг другу подходят.

    Про формат подробнее. Раньше строка формата «Статья» отбрасывалась
    целиком – Click умел только посты. Теперь у статей есть адресат: Дзен.
    Поэтому формат отсекает не пост, а ЦЕЛИ внутри него: статья уходит
    только в Дзен, обычный пост – во все площадки, включая Дзен (короткая
    публикация там тоже возможна). Видео так и остаётся вне работы: ни одна
    подключённая площадка его не примет.

    Возвращает копии постов с урезанным targets (только несформированные цели).
    """
    if today is None:
        today = apptime.now().date()
    out: list[dict] = []
    for p in posts:
        d = parse_date(p["date"])
        if d is None or d < today:
            continue
        if not (p.get("text") or "").strip():
            continue
        fmt = (p.get("format") or "Пост").strip().lower()
        if fmt.startswith("видео"):
            continue
        article = is_article(p)
        pending = [t for t in p["targets"]
                   if t["network"] in SUPPORTED
                   and not (t.get("published_link") or "").strip()
                   and (not article or t["network"] in ARTICLE_NETWORKS)]
        if pending:
            out.append({**p, "targets": pending})
    return out


# ─── Живой источник: Google-таблица (тот же аккаунт, что у КП) ───────
def _u16_index_map(base: str) -> list[int]:
    """
    Индекс UTF-16 → индекс питоновской строки (по кодовым точкам).

    Зачем. Google отдаёт границы жирного (startIndex в textFormatRuns) в
    единицах UTF-16. Эмодзи вне основного диапазона (🏗️, ⚙️, 🚀 …) в UTF-16
    занимают ДВЕ единицы, а в Python – ОДИН символ. Резать `base[start:end]`
    напрямую после такого эмодзи значило съесть по одной букве у каждого
    последующего жирного куска: «Главные» превращалось в «лавные», а Ctrl+B
    в ОК и МАКС ложился со второго символа (живой прогон 18.08.2026).

    Возвращает список: m[u16] = индекс питоновской строки для смещения u16.
    Длиннее base на 1 – последним элементом стоит конец строки.
    """
    m: list[int] = []
    for py, ch in enumerate(base):
        m.append(py)
        if ord(ch) > 0xFFFF:              # суррогатная пара – вторая единица UTF-16
            m.append(py)                  # ведёт к тому же питоновскому символу
    m.append(len(base))
    return m


def _run_link_uri(fmt: dict) -> str:
    """Адрес ссылки этого куска, если на нём в таблице стоит гиперссылка."""
    uri = ((fmt or {}).get("link") or {}).get("uri") or ""
    return uri.strip()


def _rich_runs_to_markup(runs: list[tuple[str, bool, str]]) -> str:
    """
    Куски (текст, жирный, адрес) → разметка `**[ярлык](адрес)**`.

    Пока ни на одном куске нет ссылки – отдаём ровно как раньше
    (post_text.runs_to_markup), чтобы обычные посты не менялись ни на символ.
    Как только появляется хоть одна ссылка, собираем разметку сами: слово-
    анкор из таблицы становится `[слово](адрес)`, а если оно ещё и жирное –
    `**[слово](адрес)**`. Ведущие и хвостовые пробелы выносим за маркеры,
    иначе они попадут внутрь ярлыка/жирного и «съедут» при вставке.
    """
    import post_text
    if not any(uri for _, _, uri in runs):
        return post_text.runs_to_markup([(t, b) for t, b, _ in runs])
    out: list[str] = []
    for text, bold, uri in runs:
        if not text:
            continue
        core = text.strip()
        if not core:                       # кусок из одних пробелов
            out.append(text)
            continue
        lead = text[:len(text) - len(text.lstrip())]
        trail = text[len(text.rstrip()):]
        seg = f"[{core}]({uri})" if uri else core
        if bold:
            seg = f"**{seg}**"
        out.append(lead + seg + trail)
    return "".join(out)


def _google_cell_markup(v: dict) -> str:
    """
    Ячейка из сетки Sheets API → строка с разметкой жирного И ссылок.

    Жирные куски и гиперссылки Google отдаёт одним списком textFormatRuns:
    каждый кусок начинается со startIndex и тянется до начала следующего, а
    его формат несёт и `bold`, и `link.uri`. Текст до первого куска (и ячейка
    вовсе без кусков) живёт форматом самой ячейки – effectiveFormat.textFormat.

    Ссылки читаем ОТТУДА ЖЕ, что и жирное: заказчица отмечает анкор прямо в
    реестре гиперссылкой на слове («сеткой кладочной», «нашем сайте»), и Click
    ставит ровно её – а не угадывает адрес и не тащит его голым в скобках.

    ВАЖНО: startIndex – в единицах UTF-16, а не питоновских символах. Перед
    нарезкой переводим их через _u16_index_map, иначе после эмодзи у кусков
    отваливается первая буква (см. её докстринг).
    """
    import post_text
    base = ((v.get("userEnteredValue") or {}).get("stringValue"))
    if base is None:
        base = v.get("formattedValue") or ""
    if not base:
        return ""
    cell_bold = bool(((v.get("effectiveFormat") or {}).get("textFormat") or {}).get("bold"))
    fmt_runs = v.get("textFormatRuns")
    if not fmt_runs:
        return post_text.runs_to_markup([(base, cell_bold)])

    u16 = _u16_index_map(base)

    def py_at(u16_idx: int) -> int:
        """UTF-16 смещение → индекс питоновской строки, с защитой от выхода за край."""
        if u16_idx < 0:
            return 0
        if u16_idx >= len(u16):
            return len(base)
        return u16[u16_idx]

    runs: list[tuple[str, bool, str]] = []
    first = fmt_runs[0].get("startIndex", 0)
    if first > 0:
        runs.append((base[:py_at(first)], cell_bold, ""))
    for i, run in enumerate(fmt_runs):
        start = py_at(run.get("startIndex", 0))
        end = (py_at(fmt_runs[i + 1].get("startIndex", 0))
               if i + 1 < len(fmt_runs) else len(base))
        fmt = run.get("format") or {}
        bold = fmt.get("bold")
        runs.append((base[start:end], cell_bold if bold is None else bool(bold),
                     _run_link_uri(fmt)))
    return _rich_runs_to_markup(runs)


def load_from_google(sheet_url: str, brand: str) -> list[dict]:
    """
    Прочитать лист бренда из Google-таблицы реестра через сервисный аккаунт КП.

    Отличие от чтения КП: города читаются значениями (values.get), а реестру
    нужно ФОРМАТИРОВАНИЕ – жирное из ячеек. Поэтому нативную таблицу читаем
    сеткой (spreadsheets.get с textFormatRuns), а залитый на Диск .xlsx качаем
    файлом и разбираем с rich text. Требует настроенного gcp_service_account
    (тот же, что у КП). Бросает RuntimeError с причиной словами.
    """
    import io

    import requests

    import kp_sheet
    sa = kp_sheet.service_account_info()
    if not sa:
        raise RuntimeError("Нет доступа к Google: задайте gcp_service_account (как для КП).")
    fid = kp_sheet.sheet_id(sheet_url)
    if not fid:
        raise RuntimeError("Не разобрал ссылку на таблицу реестра.")
    headers = kp_sheet._auth_headers(sa)

    # Залитый Excel: Sheets API его не читает – качаем файл и разбираем сами.
    if kp_sheet._file_mime(fid, headers) not in (kp_sheet.GOOGLE_SHEET_MIME, ""):
        r = requests.get(f"https://www.googleapis.com/drive/v3/files/{fid}",
                         params={"alt": "media", "supportsAllDrives": "true"},
                         headers=headers, timeout=90)
        if r.status_code != 200:
            raise RuntimeError(f"Drive API вернул HTTP {r.status_code} "
                               "(расшарена ли таблица на сервисный аккаунт?)")
        return parse_workbook_bytes(io.BytesIO(r.content), brand)

    base = f"https://sheets.googleapis.com/v4/spreadsheets/{fid}"
    meta = requests.get(base, params={"fields": "sheets.properties(title)"},
                        headers=headers, timeout=30)
    if meta.status_code == 403:
        raise RuntimeError("Google отказал (403). Расшарена ли таблица реестра на "
                           f"{sa.get('client_email', 'сервисный аккаунт')} как Читатель?")
    if meta.status_code != 200:
        raise RuntimeError(f"Sheets API вернул HTTP {meta.status_code}: {meta.text[:160]}")
    titles = [p.get("properties", {}).get("title", "")
              for p in (meta.json() or {}).get("sheets", [])]
    title = _pick_brand_sheet(titles, brand)
    if not title:
        raise RuntimeError(f"В таблице нет листа для бренда {brand}. Листы: {', '.join(titles)}")

    grid = requests.get(base, params={
        "ranges": f"'{title}'",
        # textFormatRuns несёт и жирное, и ссылку (format.link.uri) – просим
        # оба явно, чтобы анкоры из реестра точно приезжали, а не «иногда».
        "fields": ("sheets.data.rowData.values("
                   "formattedValue,userEnteredValue.stringValue,"
                   "textFormatRuns(startIndex,format(bold,link)),"
                   "effectiveFormat.textFormat.bold)"),
    }, headers=headers, timeout=90)
    if grid.status_code != 200:
        raise RuntimeError(f"Не удалось прочитать лист «{title}»: HTTP {grid.status_code}")
    row_data = (((grid.json() or {}).get("sheets") or [{}])[0].get("data") or [{}])[0].get("rowData") or []
    rows = [[_google_cell_markup(v) for v in (r.get("values") or [])] for r in row_data]
    return parse_sheet(rows, brand)


def _xlsx_cell_markup(cell) -> str:
    """
    Ячейка .xlsx → строка. Жирное превращается в разметку **…** – заказчик
    выделяет жирным прямо в реестре, и это выделение должно доехать до
    Телеграма и МАКСа, а не потеряться. Два случая: жирные КУСКИ (rich text)
    и жирная ЦЕЛАЯ ячейка (жирный весь шрифт ячейки – коротких постов таких
    десятки). Служебным колонкам маркеры не мешают: parse_sheet снимает их
    везде, кроме текста поста.
    """
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    import post_text
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, CellRichText):
        runs: list[tuple[str, bool]] = []
        for block in value:
            if isinstance(block, TextBlock):
                runs.append((str(block.text), bool(block.font and block.font.b)))
            else:
                runs.append((str(block), False))
        return post_text.runs_to_markup(runs)
    s = str(value)
    if s.strip() and cell.font is not None and cell.font.b:
        return post_text.runs_to_markup([(s, True)])
    return s


def _parse_xlsx(wb, brand: str) -> list[dict]:
    title = _pick_brand_sheet(wb.sheetnames, brand)
    if not title:
        raise RuntimeError(f"В файле нет листа для бренда {brand}. Листы: {', '.join(wb.sheetnames)}")
    ws = wb[title]
    rows = [[_xlsx_cell_markup(c) for c in row] for row in ws.iter_rows()]
    return parse_sheet(rows, brand)


def parse_workbook_bytes(stream, brand: str) -> list[dict]:
    """
    Прочитать лист бренда из уже загруженного файла (BytesIO из «Кросспостинга»).

    Нужен, чтобы посмотреть свой план сразу, не дожидаясь доступа к Google:
    заказчик выгружает реестр в .xlsx и загружает файл во вкладке. Файл нигде
    не сохраняется – живёт только в текущей сессии.

    rich_text=True несовместим с read_only – читаем целиком; реестр весит
    единицы мегабайт, это секунды.
    """
    from openpyxl import load_workbook
    wb = load_workbook(stream, data_only=True, rich_text=True)
    try:
        return _parse_xlsx(wb, brand)
    finally:
        wb.close()


def load_from_xlsx(path: str, brand: str) -> list[dict]:
    """Прочитать лист бренда из .xlsx на диске – для локальной проверки."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, rich_text=True)
    try:
        return _parse_xlsx(wb, brand)
    finally:
        wb.close()


def _pick_brand_sheet(titles: list[str], brand: str) -> str:
    """Название листа для бренда: сначала точное соответствие, потом по коду."""
    for t in titles:
        if SHEET_TO_BRAND.get(t.strip()) == brand:
            return t
    for t in titles:
        if t.strip().upper() == brand:
            return t
    return ""
