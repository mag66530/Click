"""
tests_audit.py – самопроверка сверки КП с организациями Яндекса.

Запуск:  python tests_audit.py

Браузер не нужен: сюда приходят уже прочитанные строки таблицы и уже
собранный список организаций, а модуль kp_audit – чистая логика. Проверяем то,
что ломается в жизни:

  • «городской округ Шахты, Шахты» и город «Шахты» – это один город;
  • «Ростовская область» городом не считается, иначе она перетянет карточку;
  • телефон «7 (385) 225-26-58» и «+7 (3852) 25-26-58» – один номер;
  • два города с похожими названиями не растаскивают карточки друг друга;
  • дубли в одном городе видны отдельной пометкой;
  • выгрузка отдаёт исходную таблицу без изменений плюс новые колонки.

Плюс разбор настоящих страниц Яндекса из tests_fixtures: если Яндекс поменяет
формат, тест это покажет раньше заказчика.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import kp_audit as A  # noqa: E402

FIXTURES = Path(__file__).parent / "tests_fixtures"
FAILED: list[str] = []
PASSED = 0


def check(name: str, condition: bool, detail: str = "") -> None:
    global PASSED
    if condition:
        PASSED += 1
        print(f"  ✅ {name}")
    else:
        FAILED.append(name)
        print(f"  ❌ {name}" + (f" – {detail}" if detail else ""))


def eq(name: str, got, expected) -> None:
    check(name, got == expected, f"получили {got!r}, ждали {expected!r}")


# ─── общая заготовка КП ─────────────────────────────────────────────
KP_HEADER = ["Страна", "Город", "URL", "Адрес", "Почта", "Общий Город", "Аккаунт", "Статус"]


def kp(*rows: list[str]) -> list[list[str]]:
    return [["", "", "", "", "", "", "Яндекс Бизнес", ""], list(KP_HEADER), *[list(r) for r in rows]]


def company(cid: str, address: str, site: str = "", phones=(), emails=(), **kw) -> dict:
    # tycoonId есть у всего, что собрано начиная с этой версии: в КП ссылки
    # записаны старым номером карточки, и без него сверка ругалась зря.
    return {"id": cid, "tycoonId": kw.get("tycoonId", ""),
            "permanentId": kw.get("permanentId", cid),
            "type": kw.get("type", "ordinal"),
            "publishing": kw.get("publishing", "publish"),
            "name": kw.get("name", "Авиапромсталь"), "names": [kw.get("name", "Авиапромсталь")],
            "address": address, "site": site, "sites": [site] if site else [],
            "social": kw.get("social", {}), "emails": list(emails), "phones": list(phones),
            "rubrics": kw.get("rubrics", []), "noAccess": kw.get("noAccess", False)}


# ════════════════════════════════════════════════════════════════════
def test_normalize() -> None:
    print("\n▸ Названия, адреса, телефоны")

    eq("городской округ снимается", A.norm_city("городской округ Шахты"), "шахты")
    eq("«г.» снимается", A.norm_city("г. Барнаул"), "барнаул")
    eq("ё сводится к е", A.norm_city("Артём"), "артем")
    eq("скобки отбрасываются", A.norm_city("Шахты (склад)"), "шахты")
    eq("дефис единый", A.norm_city("Ростов‑на‑Дону"), "ростов-на-дону")

    places = A.address_places("Южный федеральный округ, Ростовская область, "
                              "городской округ Шахты, Шахты")
    eq("из адреса остаётся только город", places, {"шахты"})
    check("область городом не считается", "ростовская область" not in places)
    check("улица городом не считается",
          "улица апатова" not in A.address_places("городской округ Мариуполь, Мариуполь, улица Апатова, 116А"))
    eq("город из короткого адреса", A.address_places("Баку, улица Башира Сафароглы, 123"), {"баку"})

    eq("хост без схемы и слэша", A.site_host("https://Shahty.Aviastal.RU/"), "shahty.aviastal.ru")
    eq("хост без www", A.site_host("www.aviastal.ru/page"), "aviastal.ru")
    eq("пустая ссылка", A.site_host(""), "")

    eq("телефон: код страны не мешает",
       A.phone_key("7 (391) 271-75-19"), A.phone_key("+7 (391) 271-75-19"))
    eq("телефон: 8 и 7 – один номер",
       A.phone_key("8 (800) 700-36-89"), A.phone_key("+7 (800) 700-36-89"))
    eq("телефон: скобки в разных местах",
       A.phone_key("7 (385) 225-26-58"), A.phone_key("+7 (3852) 25-26-58"))
    eq("не телефон", A.phone_key("нет"), "")


def test_parse_sheet() -> None:
    print("\n▸ Разбор листа КП")

    rows = kp(["Россия", "Абакан", "https://abakan.aviastal.ru", "ул. Щетинкина, 24",
               "abakan@aviastal.ru", "7 (391) 271-75-19", "", ""],
              ["Россия", "Барнаул", "barnaul.aviastal.ru", "ул. Пушкина, 29",
               "barnaul@aviastal.ru", "7 (385) 225-26-58",
               "https://yandex.ru/sprav/21461411/edit/main", "Активная"])
    got = A.parse_sheet(rows)
    eq("шапка найдена", got["headerIdx"], 1)
    eq("городов", len(got["items"]), 2)
    eq("колонка города", got["columns"]["city"], 1)
    eq("колонка ссылки – та, где ссылки на sprav", got["columns"]["link"], 6)
    eq("номер строки сохранён", got["items"][1]["rowIdx"], 3)
    eq("почта прочитана", got["items"][0]["email"], "abakan@aviastal.ru")
    eq("телефон прочитан", got["items"][1]["phones"], ["7 (385) 225-26-58"])

    # Колонка сайта без заголовка – ровно как в КП АПС.
    blind = [["", "", "", ""], ["Страна", "Город", "", "Почта"],
             ["Россия", "Абакан", "https://abakan.aviastal.ru", "abakan@aviastal.ru"],
             ["Россия", "Бийск", "https://bijsk.aviastal.ru", "bijsk@aviastal.ru"],
             ["Россия", "Псков", "https://pskov.aviastal.ru", "pskov@aviastal.ru"]]
    eq("сайт найден по содержимому", A.parse_sheet(blind)["columns"]["site"], 2)

    eq("без шапки – понятная ошибка",
       A.parse_sheet([["что-то"], ["ещё"]])["error"],
       "Не нашли шапку: нужна строка с колонкой «Город»")


def test_match() -> None:
    print("\n▸ Сопоставление городов и карточек")

    rows = kp(["Россия", "Шахты", "https://shahty.aviastal.ru", "", "", "", "", ""],
              ["Россия", "Ростов-на-Дону", "https://rostov.aviastal.ru", "", "", "", "", ""],
              ["Россия", "Бердск", "https://berdsk.aviastal.ru", "", "", "", "", ""])
    cos = [
        company("1", "Южный федеральный округ, Ростовская область, городской округ Шахты, Шахты",
                "https://shahty.aviastal.ru/"),
        company("2", "Южный федеральный округ, Ростовская область, Ростов-на-Дону, улица Ленина, 1",
                "https://rostov.aviastal.ru/"),
        company("3", "Сибирский федеральный округ, Новосибирская область, городской округ Бердск, Бердск",
                "https://berdsk.aviastal.ru/"),
    ]
    res = A.build(rows, cos)
    by_city = {i["city"]: [c["id"] for c in i["companies"]] for i in res["items"]}
    eq("Шахты – своя карточка", by_city["Шахты"], ["1"])
    eq("Ростов-на-Дону – своя", by_city["Ростов-на-Дону"], ["2"])
    eq("Бердск – своя", by_city["Бердск"], ["3"])
    eq("лишних нет", res["extra"], [])
    eq("область не перетянула карточку", res["totals"]["missing"], 0)

    # Карточка без сайта, адрес называет город – всё равно находится.
    res2 = A.build(kp(["Россия", "Псков", "", "", "", "", "", ""]),
                   [company("9", "Северо-Западный федеральный округ, Псковская область, "
                                 "городской округ Псков, Псков")])
    eq("нашли по одному адресу", [c["id"] for c in res2["items"][0]["companies"]], ["9"])

    # Ничего похожего – организация уходит в «лишние», город остаётся пустым.
    res3 = A.build(kp(["Россия", "Псков", "https://pskov.aviastal.ru", "", "", "", "", ""]),
                   [company("7", "Марс, кратер Гусева", "https://mars.aviastal.ru/")])
    eq("чужая карточка – в лишние", [c["id"] for c in res3["extra"]], ["7"])
    eq("город остался без карточки", res3["items"][0]["cmp"]["status"], "нет")

    # Сетевая карточка города не имеет – её в города не пихаем.
    res4 = A.build(kp(["Россия", "Псков", "", "", "", "", "", ""]),
                   [company("5", "Земля", "https://aviastal.kz/", type="chain")])
    eq("сеть отдельно", [c["id"] for c in res4["chains"]], ["5"])
    eq("сеть не села в город", res4["items"][0]["cmp"]["status"], "нет")

    # Ссылка в КП сильнее всего: адрес пустой, а карточка та самая.
    res5 = A.build(kp(["Россия", "Неведомск", "", "", "", "",
                       "https://yandex.ru/sprav/4242/edit/main", ""]),
                   [company("4242", "Где-то там")])
    eq("нашли по ссылке из КП", [c["id"] for c in res5["items"][0]["companies"]], ["4242"])
    eq("сопоставили по ссылке", res5["items"][0]["companies"][0]["matchedBy"], "ссылка из КП")


def test_same_name_cities() -> None:
    """
    Города-тёзки и города с двумя именами. Оба случая есть в живом «Лист20»:
    Армавир стоит и в России, и в Армении, а Астана записана как
    «Астана/Нур-Султан» – Яндекс знает только второе имя.
    """
    print("\n▸ Тёзки и вторые названия городов")

    eq("косая черта – два написания",
       A.city_variants("Астана/Нур-Султан"), ["астана", "нур-султан"])
    eq("пробел перед чертой не мешает",
       A.city_variants("Нахичевань /Нахчыван"), ["нахичевань", "нахчыван"])
    eq("страна по названию", A.country_code("Армения"), "AM")
    eq("РФ и Россия – одна страна", A.country_code("РФ"), A.country_code("Россия"))
    eq("незнакомая страна", A.country_code("Атлантида"), "")

    rows = kp(["Россия", "Армавир", "https://armavir.aviastal.ru", "", "", "", "", ""],
              ["Армения", "Армавир", "armavir.aviastal.am", "", "", "", "", ""],
              ["Казахстан", "Астана/Нур-Султан", "", "", "", "", "", ""],
              ["Беларусь", "Минск", "", "", "", "", "", ""])
    ru = company("RU1", "Краснодарский край, городской округ Армавир, Армавир")
    ru["regionCode"] = "RU"
    am = company("AM1", "Армавир, улица Шаумяна, 1")
    am["regionCode"] = "AM"
    kz = company("KZ1", "Нур-Султан, проспект Мангилик Ел, 10")
    kz["regionCode"] = "KZ"
    stray = company("X1", "Минск, проспект Независимости, 1")
    stray["regionCode"] = "RU"          # у карточки страна не та, что в адресе

    res = A.build(rows, [ru, am, kz, stray])
    got = {f'{i["city"]}|{i["country"]}': [c["id"] for c in i["companies"]] for i in res["items"]}
    eq("российский Армавир взял свою", got["Армавир|Россия"], ["RU1"])
    eq("армянский Армавир взял свою", got["Армавир|Армения"], ["AM1"])
    eq("Астана нашлась по второму имени", got["Астана/Нур-Султан|Казахстан"], ["KZ1"])
    eq("карточка из другой страны в город не села", got["Минск|Беларусь"], [])
    eq("она ушла в лишние", [c["id"] for c in res["extra"]], ["X1"])

    # Одной страны мало: без города и сайта карточка не должна садиться никуда.
    lone = company("L1", "Где-то в России")
    lone["regionCode"] = "RU"
    only_country = A.build(kp(["Россия", "Тула", "", "", "", "", "", ""]), [lone])
    eq("совпадения одной страны недостаточно", [c["id"] for c in only_country["extra"]], ["L1"])

    # Похожие названия: одно – начало другого.
    pair = kp(["Россия", "Петропавловск-Камчатский", "", "", "", "", "", ""],
              ["Казахстан", "Петропавловск", "", "", "", "", "", ""])
    p1 = company("P1", "Камчатский край, Петропавловск-Камчатский")
    p1["regionCode"] = "RU"
    p2 = company("P2", "Северо-Казахстанская область, Петропавловск")
    p2["regionCode"] = "KZ"
    res2 = A.build(pair, [p1, p2])
    got2 = {i["city"]: [c["id"] for c in i["companies"]] for i in res2["items"]}
    eq("Камчатский взял свою", got2["Петропавловск-Камчатский"], ["P1"])
    eq("казахский взял свою", got2["Петропавловск"], ["P2"])


def test_duplicates() -> None:
    print("\n▸ Несколько карточек в одном городе")

    rows = kp(["Россия", "Шахты", "https://shahty.aviastal.ru", "", "", "", "", ""])
    cos = [company("1", "Ростовская область, городской округ Шахты, Шахты", "https://shahty.aviastal.ru/"),
           company("2", "Ростовская область, городской округ Шахты, Шахты, улица Ленина, 5")]
    res = A.build(rows, cos)
    item = res["items"][0]
    eq("обе карточки в городе", sorted(c["id"] for c in item["companies"]), ["1", "2"])
    eq("город помечен", item["cmp"]["status"], "несколько")
    check("в пометке сказано про дубли",
          any("дубли" in n for n in item["cmp"]["problems"]),
          str(item["cmp"]["problems"]))
    eq("счётчик дублей", res["totals"]["several"], 1)


def test_compare() -> None:
    print("\n▸ Сравнение полей")

    rows = kp(["Россия", "Шахты", "https://shahty.aviastal.ru", "", "shahty@aviastal.ru",
               "7 (863) 206-68-85", "https://yandex.ru/sprav/1/edit/", ""])
    same = A.build(rows, [company("1", "Ростовская область, Шахты", "https://shahty.aviastal.ru/",
                                  phones=["+7 (863) 206-68-85"], emails=["shahty@aviastal.ru"])])
    cmp = same["items"][0]["cmp"]
    eq("сайт совпал", cmp["site"], "совпадает")
    eq("телефон совпал", cmp["phone"], "совпадает")
    eq("почта совпала", cmp["email"], "совпадает")
    eq("разногласий нет", cmp["problems"], [])
    check("строка считается чистой", cmp["ok"])
    eq("счётчик чистых", same["totals"]["clean"], 1)

    other = A.build(rows, [company("1", "Ростовская область, Шахты", "https://old-shahty.ru/",
                                   phones=["+7 (999) 000-00-00"], emails=[])])
    cmp2 = other["items"][0]["cmp"]
    eq("сайт разошёлся", cmp2["site"], "расходится")
    eq("телефон разошёлся", cmp2["phone"], "расходится")
    eq("почты нет в Яндексе", cmp2["email"], "нет в Яндексе")
    check("строка попала в расхождения", not cmp2["ok"])
    eq("счётчик расхождений", other["totals"]["mismatch"], 1)

    # Телефонов у Яндекса больше – номер из КП среди них, это совпадение.
    more = A.build(rows, [company("1", "Ростовская область, Шахты", "https://shahty.aviastal.ru/",
                                  phones=["+7 (863) 206-68-85", "+7 (999) 111-22-33"],
                                  emails=["shahty@aviastal.ru"])])
    eq("лишний номер не ломает совпадение", more["items"][0]["cmp"]["phone"], "совпадает")

    # Ссылка в КП ведёт не туда – это разногласие, а не «нашли». Карточки 999
    # среди прочитанных нет, поэтому честный диагноз – устаревший список,
    # а не «ссылка на другую карточку»: см. проверки ниже про all_ids.
    wrong = A.build(kp(["Россия", "Шахты", "https://shahty.aviastal.ru", "", "", "",
                        "https://yandex.ru/sprav/999/edit/", ""]),
                    [company("1", "Ростовская область, Шахты", "https://shahty.aviastal.ru/")])
    check("несовпадение по ссылке замечено",
          any("999" in n for n in wrong["items"][0]["cmp"]["problems"]),
          str(wrong["items"][0]["cmp"]["problems"]))

    # Старый номер карточки. В КП ссылки записаны им, список организаций
    # отдаёт новый – это ОДНА карточка, ругаться не на что.
    old_link = kp(["Россия", "Шахты", "https://shahty.aviastal.ru", "", "", "",
                   "https://yandex.ru/sprav/25755702/edit/main", ""])
    same_card = company("10873194809", "Ростовская область, Шахты",
                        "https://shahty.aviastal.ru/", tycoonId="25755702")
    old = A.build(old_link, [same_card])
    eq("старый номер карточки – та же карточка", old["items"][0]["cmp"]["problems"], [])
    eq("и сопоставили именно по ссылке",
       old["items"][0]["companies"][0]["matchedBy"], "ссылка из КП")

    # Данные, собранные до появления tycoonId, сравнивать нечем – молчим.
    legacy = dict(same_card)
    legacy.pop("tycoonId")
    legacy["id"] = "999999"
    quiet = A.build(old_link, [legacy])
    check("на старых данных ложной тревоги нет",
          not any("другую карточку" in n for n in quiet["items"][0]["cmp"]["problems"]),
          str(quiet["items"][0]["cmp"]["problems"]))

    # Пустая ссылка в КП – это подсказка, а не ошибка: её сверка и заполняет.
    empty = A.build(kp(["Россия", "Шахты", "https://shahty.aviastal.ru", "", "", "", "", ""]),
                    [company("1", "Ростовская область, Шахты", "https://shahty.aviastal.ru/")])
    eq("пустая ссылка – не ошибка", empty["items"][0]["cmp"]["problems"], [])
    check("но подсказка есть", bool(empty["items"][0]["cmp"]["hints"]))
    eq("счётчик «без ссылки»", empty["totals"]["noLink"], 1)


def test_export() -> None:
    print("\n▸ Выгрузка: КП с колонками из Яндекса")

    rows = kp(["Россия", "Шахты", "https://shahty.aviastal.ru", "ул. Ленина, 1",
               "shahty@aviastal.ru", "7 (863) 206-68-85", "", ""],
              ["Россия", "Мытищи", "https://mytishchi.aviastal.ru", "", "", "", "", ""])
    res = A.build(rows, [company("1", "Ростовская область, Шахты", "https://shahty.aviastal.ru/",
                                 phones=["+7 (863) 206-68-85"], emails=["shahty@aviastal.ru"])])
    out = A.to_rows(rows, res)
    col = {name: len(KP_HEADER) + n for n, name in enumerate(A.EXTRA_HEADERS)}

    eq("строк столько же", len(out), len(rows))
    for i, src in enumerate(rows):
        eq(f"строка {i} не тронута", out[i][:len(src)], [str(c) for c in src])
    eq("шапка получила новые колонки", out[1][len(KP_HEADER):], A.EXTRA_HEADERS)
    eq("статус найденного города", out[2][col["Проверка"]], "✅ найдена")
    eq("статус ненайденного", out[3][col["Проверка"]], "❌ нет")
    check("ссылка на карточку подставлена",
          "yandex.ru/sprav/1/p/edit/" in out[2][col["Ссылка на карточку"]],
          out[2][col["Ссылка на карточку"]])
    eq("вывод по сайту рядом со статусом", out[2][col["Сайт совпадает?"]], "совпадает")

    csv = A.to_csv(out)
    eq("в CSV столько же строк", len(csv.strip().splitlines()), len(rows))
    check("перенос строки внутри ячейки не рвёт CSV", "\n" not in csv.strip().splitlines()[2])


def test_report_sheets() -> None:
    """
    Отчёт книгой: у каждого вопроса свой лист. Заказчик про прошлый вариант
    сказала «ничего не понятно» – одна таблица в сорок колонок отвечала сразу
    на всё.
    """
    print("\n▸ Отчёт книгой")

    rows = kp(["Россия", "Шахты", "https://shahty.aviastal.ru", "", "shahty@aviastal.ru",
               "7 (863) 206-68-85", "", ""],
              ["Россия", "Мытищи", "https://mytishchi.aviastal.ru", "", "", "", "", ""],
              ["Россия", "Казань", "https://kazan.aviastal.ru", "", "", "7 (843) 216-39-74", "", ""],
              ["Россия", "Псков", "https://pskov.aviastal.ru", "", "pskov@aviastal.ru", "", "", ""])
    cos = [
        # Шахты: всё сходится
        company("1", "Ростовская область, городской округ Шахты, Шахты",
                "https://shahty.aviastal.ru/", phones=["+7 (863) 206-68-85"],
                emails=["shahty@aviastal.ru"]),
        # Казань: два дубля
        company("2", "Республика Татарстан, Казань", "https://kazan.aviastal.ru/",
                phones=["+7 (843) 216-39-74"]),
        company("3", "Республика Татарстан, Казань, улица Баумана, 1"),
        # Псков: сайт и почта расходятся
        company("4", "Псковская область, городской округ Псков, Псков",
                "https://old-pskov.ru/", emails=["other@aviastal.ru"]),
        # Ярославль: в КП такого города нет
        company("5", "Центральный федеральный округ, Ярославская область, "
                     "городской округ Ярославль, Ярославль", "https://yaroslavl.aviastal.ru/",
                phones=["+7 (4852) 66-29-46"], emails=["yaroslavl@aviastal.ru"]),
    ]
    res = A.build(rows, cos)

    # ── списки под конкретный вопрос ──
    extra = A.extra_rows(res)
    eq("«Нет в КП»: одна организация", len(extra) - 1, 1)
    eq("город подставлен из адреса", extra[1][0], "Ярославль")
    check("и ссылка на карточку рядом", extra[1][-1].endswith("/sprav/5/p/edit/"), extra[1][-1])

    doubles = A.double_rows(res)
    eq("«Дубли»: две строки на один город", len(doubles) - 1, 2)
    eq("обе про Казань", {r[1] for r in doubles[1:]}, {"Казань"})
    eq("карточки пронумерованы", [r[2] for r in doubles[1:]], ["1", "2"])

    # Ссылка из КП: карточка в аккаунте есть, но привязана к другому городу –
    # это неверная ссылка. А если карточки нет среди прочитанных вовсе, дело
    # не в ссылке: список организаций устарел. Заказчик наткнулась именно на
    # второе – в Яндексе у города две карточки, в отчёте одна, и «ссылка
    # ведёт на другую» звучало как неправда.
    карточка = {"id": "777", "tycoonId": "776", "type": "ordinal", "publishing": "publish",
                "name": "МПИ", "address": "Красноярск", "site": "", "sites": [],
                "emails": [], "phones": [], "social": {}, "rubrics": [], "noAccess": False}
    строка = {"country": "Россия", "city": "Красноярск", "site": "", "email": "",
              "phones": [], "link": "https://yandex.ru/sprav/555/", "rowIdx": 1}
    чужая = A.compare(строка, [карточка], all_ids={"555", "777", "776"})["problems"]
    check("карточка по ссылке есть в аккаунте – это неверная ссылка",
          any("ведёт на другую карточку" in p for p in чужая), str(чужая))
    нет = A.compare(строка, [карточка], all_ids={"777", "776"})["problems"]
    check("карточки по ссылке нет среди прочитанных – сказано про устаревший список",
          any("нет среди прочитанных" in p for p in нет), str(нет))
    check("и подсказано, что делать", any("перечитайте список" in p for p in нет), str(нет))

    # Две карточки на один город – это дубль, и он обязан попасть на лист.
    двое = [
        {"id": "1", "type": "ordinal", "publishing": "publish", "name": "МПИ",
         "address": "Красноярский край, городской округ Красноярск, Красноярск, "
                    "Тамбовская улица, 5", "site": "", "sites": [], "emails": [],
         "phones": [], "social": {}, "rubrics": [], "noAccess": False, "regionCode": "RU"},
        {"id": "2", "type": "ordinal", "publishing": "publish", "name": "МПИ",
         "address": "Красноярск", "site": "https://krasnoyarsk.mpi.ru/",
         "sites": ["https://krasnoyarsk.mpi.ru/"], "emails": [], "phones": [],
         "social": {}, "rubrics": [], "noAccess": False, "regionCode": "RU"},
    ]
    пара = A.build([["Страна", "Город", "Сайт", "", "", "", "", ""],
                    ["Россия", "Красноярск", "https://krasnoyarsk.mpi.ru/", "", "", "", "", ""]],
                   двое)
    eq("две карточки на город – статус «несколько»",
       (пара["items"][0]["cmp"] or {}).get("status"), "несколько")
    eq("обе попали на лист «Дубли»", len(A.double_rows(пара)) - 1, 2)
    eq("и ни одна не считается лишней", len(пара["extra"]), 0)

    # Заказчик: «дубли – это же ещё и два Красноярска в Организациях». Ловим
    # их в любом виде: обе карточки сели на город КП, одна села а вторая нет,
    # и города в КП вовсе нет.
    ADR1 = ("Сибирский федеральный округ, Красноярский край, городской округ "
            "Красноярск, Красноярск, Тамбовская улица, 5, корп. 2, стр. 1")
    ADR2 = "Красноярский край, городской округ Красноярск, Красноярск, улица Ленина, 1"
    ADR3 = "Красноярский край, городской округ Красноярск, Красноярск, улица Мира, 7"

    def кп_город(город: str, лат: str):
        return [["Страна", "Город", "Сайт", "", "", "", "", ""],
                ["Россия", город, f"https://{лат}.mpi.ru/", "", "", "", "", ""]]

    def карточка(cid: str, адрес: str, сайт: str = ""):
        return {"id": cid, "type": "ordinal", "publishing": "publish", "name": "МПИ",
                "address": адрес, "site": сайт, "sites": [сайт] if сайт else [],
                "emails": [], "phones": [], "social": {}, "rubrics": [],
                "noAccess": False, "regionCode": "RU"}

    обе = A.build(кп_город("Красноярск", "krasnoyarsk"),
                  [карточка("1", ADR1, "https://krasnoyarsk.mpi.ru/"), карточка("2", ADR2)])
    eq("две карточки на город КП – «несколько»",
       (обе["items"][0]["cmp"] or {}).get("status"), "несколько")
    eq("и обе на листе «Дубли»", len(A.double_rows(обе)) - 1, 2)

    # Карточки могли сесть на РАЗНЫЕ строки КП – город всё равно один.
    две_строки = A.build([["Страна", "Город", "Сайт", "", "", "", "", ""],
                          ["Россия", "Красноярск", "https://krasnoyarsk.mpi.ru/", "", "", "", "", ""],
                          ["Россия", "Ачинск", "https://achinsk.mpi.ru/", "", "", "", "", ""]],
                         [карточка("1", ADR1), карточка("2", "Красноярск")])
    eq("две карточки Красноярска на разных строках КП – всё равно дубль",
       len(A.double_rows(две_строки)) - 1, 2)

    # Карточка «онлайн»: адрес – одно название города, без улицы. Ровно она
    # и висит вторым Красноярском в аккаунте заказчика.
    онлайн = A.build(кп_город("Красноярск", "krasnoyarsk"),
                     [карточка("1", ADR1),
                      карточка("2", "Красноярск", "https://krasnoyarsk.mpi.ru/")])
    eq("голый «Красноярск» в адресе – тот же город",
       len(A.double_rows(онлайн)) - 1, 2)

    нет_города = A.build(кп_город("Омск", "omsk"),
                         [карточка("1", ADR1), карточка("2", ADR2)])
    строки = A.double_rows(нет_города)[1:]
    eq("два Красноярска, которых в КП нет, – тоже дубль", len(строки), 2)
    eq("город определён по адресу", {r[1] for r in строки}, {"Красноярск"})
    check("и помечено, что города нет в КП",
          all("нет в КП" in r[0] for r in строки), str([r[0] for r in строки]))

    три = A.build(кп_город("Красноярск", "krasnoyarsk"),
                  [карточка("1", ADR1), карточка("2", ADR2), карточка("3", ADR3)])
    eq("три карточки – три строки", len(A.double_rows(три)) - 1, 3)
    eq("пронумерованы", [r[2] for r in A.double_rows(три)[1:]], ["1", "2", "3"])

    один = A.build(кп_город("Омск", "omsk"),
                   [карточка("1", "Омская область, городской округ Омск, Омск")])
    eq("одна карточка – дублей нет", len(A.double_rows(один)) - 1, 0)

    # Ложных дублей быть не должно: разные города в один не сливаются.
    разные = A.build([["Страна", "Город", "Сайт", "", "", "", "", ""],
                      *[["Россия", г, "", "", "", "", "", ""]
                        for г in ("Омск", "Томск", "Тюмень", "Курск")]],
                     [карточка("1", "Омская область, городской округ Омск, Омск, Ленина, 1"),
                      карточка("2", "Томская область, городской округ Томск, Томск, Ленина, 2"),
                      карточка("3", "Тюменская область, городской округ Тюмень, Тюмень, Мира, 3"),
                      карточка("4", "Курская область, городской округ Курск, Курск, Мира, 4")])
    eq("одинаковые улицы в разных городах – не дубль", len(A.double_rows(разные)) - 1, 0)

    # Сетевая карточка одна на все города – она не дубль ничему.
    сеть = карточка("s", "Земля")
    сеть["type"] = "chain"
    с_сетью = A.build(кп_город("Омск", "omsk"),
                      [карточка("1", "Омская область, городской округ Омск, Омск"), сеть])
    eq("сеть не считается дублем", len(A.double_rows(с_сетью)) - 1, 0)

    # Улица и номер дома городом не считаются. На живом адресе из Красноярска
    # «городом» выходило «стр. 1»: Яндекс пишет «Тамбовская улица» (слово
    # в конце), а «корп. 2» не отсекалось из-за точки в сокращении.
    eq("город из живого адреса", A.guess_city(ADR1), "Красноярск")
    eq("и в местах адреса только он", A.address_places(ADR1), {"красноярск"})
    # Казахстан: «городской акимат Актау» – та же приставка, что «городской
    # округ». Без неё повтор не находился, и городом становилась промзона
    # или номер дома. Взято с живых адресов заказчика.
    eq("акимат снимается как приставка",
       A.guess_city("Мангистауская область, городской акимат Актау, Актау, "
                    "Промышленная зона № 1, 33/3"), "Актау")
    eq("номер дома со слэшем городом не считается",
       A.guess_city("Северо-Казахстанская область, городской акимат Петропавловск, "
                    "Петропавловск, улица Габита Мусрепова, 34/В"), "Петропавловск")
    eq("сельский округ – это не город",
       A.guess_city("Западно-Казахстанская область, Теректинский район, "
                    "Акжаикский сельский округ, село Абай"), "Абай")
    eq("посёлок внутри округа – берём посёлок",
       A.guess_city("Московская область, Одинцовский городской округ, посёлок Заречье, "
                    "Заречье, Заречная улица, 2"), "Заречье")

    missing = A.missing_rows(res)
    eq("«Нет в Яндексе»: один город", len(missing) - 1, 1)
    eq("это Мытищи", missing[1][1], "Мытищи")

    # Города, где карточка есть, а ссылки на неё в КП не записано. Цифра на
    # дашборде была, а списка за ней не было: заказчик видела пометку
    # «нет ссылки в КП» в строках и не понимала, где их собрать.
    nolink = A.nolink_rows(res)
    eq("«Нет ссылки в КП»: три города", len(nolink) - 1, 3)
    eq("те, у кого карточка нашлась", {r[1] for r in nolink[1:]},
       {"Шахты", "Казань", "Псков"})
    check("Мытищи сюда не попали – карточки нет вовсе",
          "Мытищи" not in {r[1] for r in nolink[1:]})
    check("ссылка на карточку рядом",
          all(r[-1].endswith("/p/edit/") for r in nolink[1:]),
          str([r[-1] for r in nolink[1:]]))
    kazan = [r for r in nolink[1:] if r[1] == "Казань"][0]
    check("у города с дублями сказано, что карточек несколько",
          "карточек: 2" in kazan[2], kazan[2])
    eq("число на дашборде и длина списка сходятся",
       res["totals"]["noLink"], len(nolink) - 1)

    diffs = A.diff_rows(res)
    fields = {r[2].split(":")[0] for r in diffs[1:]}
    check("«Расхождения»: попали сайт и почта Пскова", {"Сайт", "Почта"} <= fields, str(fields))
    pskov = [r for r in diffs[1:] if r[1] == "Псков" and r[2].startswith("Сайт")][0]
    eq("слева КП", pskov[3], "https://pskov.aviastal.ru")
    eq("справа Яндекс", pskov[4], "https://old-pskov.ru/")

    eq("город из адреса", A.guess_city("Центральный федеральный округ, Ярославская область, "
                                       "городской округ Ярославль, Ярославль"), "Ярославль")
    eq("город из адреса с улицей",
       A.guess_city("городской округ Мариуполь, Мариуполь, улица Апатова, 116А"), "Мариуполь")

    # ── книга целиком ──
    blob = A.to_xlsx(rows, res, "Лист20", collected_at="05.08.2026 12:00")
    check("xlsx собрался", len(blob) > 5000, str(len(blob)))
    import io

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob))
    eq("листы книги", wb.sheetnames,
       ["Дашборд", "Как читать", "Нет в КП", "Нет ссылки в КП", "Дубли",
        "Нет в Яндексе", "Расхождения", "Статусы", "Сверка", "КП с данными"])

    nl = wb["Нет ссылки в КП"]
    _texts = [str(c.value or "") for row in nl.iter_rows() for c in row]
    check("на листе есть ссылка-подпись", "Открыть карточку" in _texts, "")
    check("и объяснено, что с ним делать",
          any("Скопируйте ссылку в КП" in t for t in _texts), "")

    dash = wb["Дашборд"]
    eq("на дашборде число городов", dash["B6"].value, 4)
    eq("и число ненайденных", dash["F6"].value, 1)
    check("на дашборде видно, с какого листа читали",
          "Лист20" in str(dash["B3"].value), str(dash["B3"].value))

    grid = wb["Сверка"]
    eq("сетка: строка на город", grid.max_row, 1 + 4)
    eq("шапка сетки", [c.value for c in grid[1]], A.GRID_HEADERS)
    by_city = {grid.cell(row=r, column=2).value: r for r in range(2, grid.max_row + 1)}
    eq("Шахты: карточка на месте", grid.cell(row=by_city["Шахты"], column=3).value, "✓")
    eq("Шахты: ошибок нет", grid.cell(row=by_city["Шахты"], column=9).value, 0)
    eq("Казань: дубли помечены", grid.cell(row=by_city["Казань"], column=3).value, "⚠")
    eq("Мытищи: карточки нет", grid.cell(row=by_city["Мытищи"], column=3).value, "✗")
    eq("Псков: сайт расходится", grid.cell(row=by_city["Псков"], column=4).value, "✗")
    check("и подробности висят примечанием",
          "old-pskov" in (grid.cell(row=by_city["Псков"], column=4).comment.text or ""),
          str(grid.cell(row=by_city["Псков"], column=4).comment))
    check("у зелёной ячейки заливки нет",
          not (grid.cell(row=by_city["Шахты"], column=4).fill.fgColor.rgb or "").endswith(A.C_BUG))
    check("ссылка на карточку кликабельная",
          bool(grid.cell(row=by_city["Шахты"], column=8).hyperlink))
    eq("подпись у ссылки человеческая",
       grid.cell(row=by_city["Шахты"], column=8).value, "Открыть карточку")

    full = wb["КП с данными"]
    eq("полное КП на месте: колонок", full.max_column, len(KP_HEADER) + len(A.EXTRA_HEADERS))
    eq("город не сдвинулся", full.cell(row=3, column=2).value, "Шахты")
    check("шапка закреплена", bool(full.freeze_panes), str(full.freeze_panes))
    check("фильтр включён", bool(full.auto_filter.ref), str(full.auto_filter.ref))
    wb.close()


# ─── настоящие страницы Яндекса ─────────────────────────────────────
def _preload(html: str) -> dict:
    """window.__PRELOAD_DATA из сохранённой страницы – без браузера, скобками."""
    i = html.index("window.__PRELOAD_DATA = ") + len("window.__PRELOAD_DATA = ")
    depth, instr, esc, end = 0, False, False, len(html)
    for j in range(i, len(html)):
        c = html[j]
        if instr:
            if esc:
                esc = False
            elif c == "\\":
                esc = True
            elif c == '"':
                instr = False
            continue
        if c == '"':
            instr = True
        elif c == "{":
            depth += 1
        elif c == "}":
            depth -= 1
            if depth == 0:
                end = j + 1
                break
    return json.loads(html[i:end])


def test_yandex_pages() -> None:
    print("\n▸ Настоящие страницы Яндекса")

    card_fp = FIXTURES / "yb-company-card.html"
    list_fp = FIXTURES / "yb-companies-list.html"
    if not card_fp.exists() or not list_fp.exists():
        check("фикстуры страниц Яндекса на месте", False, "нет файлов в tests_fixtures")
        return

    import yb_playwright as yb

    # Читалка живёт в браузере, поэтому здесь проверяем ДАННЫЕ, из которых она
    # берёт всё: если Яндекс уберёт __PRELOAD_DATA, сверка останется без списка.
    card = _preload(card_fp.read_text(encoding="utf-8", errors="replace"))
    co = card["initialState"]["edit"]["info"]["company"]
    eq("карточка: адрес", co["address"]["formatted"]["value"], "Баку, улица Башира Сафароглы, 123")
    eq("карточка: почта", co["emails"], ["info@aviastal.az"])
    eq("карточка: телефонов", len(co["phones"]), 2)
    eq("карточка: сайт", [u["value"] for u in co["urls"] if u["type"] == "main"], ["https://aviastal.az/"])

    lst = _preload(list_fp.read_text(encoding="utf-8", errors="replace"))["initialState"]["companiesList"]
    eq("список: страница", lst["page"], 1)
    check("список: сказано, сколько всего", lst["total"] > len(lst["listCompanies"]),
          f'{lst["total"]} vs {len(lst["listCompanies"])}')
    check("список: строки со всеми полями",
          all(k in lst["listCompanies"][0] for k in ("id", "address", "urls", "phones", "emails")))

    # Читалка списка обязана уметь дочитывать страницы: одной страницы мало.
    src = yb._COMPANIES_JS
    check("читалка списка берёт данные из preload", "__PRELOAD_DATA" in src)
    check("читалка списка отдаёт total и limit", "total:" in src and "limit:" in src)
    import inspect
    check("страницы запрашиваются параметрами page и limit",
          "page={num}" in inspect.getsource(yb._companies_page)
          and "limit={limit}" in inspect.getsource(yb._companies_page))
    check("почта в запасном разборе ищется по-русски",
          "/почт/i" in yb._CARD_DOM_JS,
          "в JS \\w не ловит кириллицу – регулярка по «электронн\\w* почт» не сработает")
    check("адрес в запасном разборе ищется по своему блоку", ".InfoAddress input" in yb._CARD_DOM_JS)

    # Сверка на настоящих данных: девять городов КП против десяти карточек.
    companies = []
    for raw in lst["listCompanies"]:
        urls = raw.get("urls") or []
        companies.append(company(
            str(raw["id"]),
            (raw.get("address") or {}).get("formatted", {}).get("value", ""),
            next((u["value"] for u in urls if u.get("type") == "main"), ""),
            phones=[p.get("formatted", "") for p in (raw.get("phones") or [])],
            emails=raw.get("emails") or [],
            type=raw.get("type", "ordinal"), name=raw.get("displayName", "")))
    real_kp = kp(*[["Россия", city, f"https://{host}.aviastal.ru", "", "", phone, "", ""]
                   for city, host, phone in [
                       ("Шахты", "shahty", "7 (863) 206-68-85"),
                       ("Бердск", "berdsk", "7 (383) 255-73-36"),
                       ("Батайск", "bataysk", "7 (863) 206-68-85"),
                       ("Саранск", "saransk", "7 (831) 414-40-86"),
                       ("Мариуполь", "mariupol", "8 (800) 700-36-89"),
                       ("Бийск", "bijsk", "7 (385) 225-26-58"),
                       ("Псков", "pskov", "7 (812) 986-43-38"),
                       ("Балашиха", "balashiha", "7 (495) 755-36-18"),
                       ("Нальчик", "nalchik", "7 (903) 411-64-68"),
                       ("Сургут", "surgut", "7 (346) 200-00-00")]])
    res = A.build(real_kp, companies)
    eq("девять городов нашлись", res["totals"]["found"], 9)
    eq("Сургута в Яндексе нет", res["totals"]["missing"], 1)
    eq("дублей нет", res["totals"]["several"], 0)
    eq("сетевая карточка отдельно", len(res["chains"]), 1)
    eq("лишних нет", len(res["extra"]), 0)
    # У Сургута карточки нет вовсе, телефон сравнивать не с чем – он не в счёт.
    bad = [i["city"] for i in res["items"]
           if i["companies"] and i["cmp"]["phone"] not in ("совпадает", "–")]
    eq("телефоны сошлись у всех найденных", bad, [])


# ════════════════════════════════════════════════════════════════════
#  Статусы: КП против площадок
# ════════════════════════════════════════════════════════════════════
#
# Шапка настоящего КП МПИ: три площадки подряд, у каждой свой блок
# «Аккаунт – Карта – Статус», и все три статуса называются одинаково –
# просто «Статус». Различает их только положение относительно ссылок.
KP2_HEAD_TOP = ["", "", "", "Яндекс Бизнес", "", "", "2ГИС", "", ""]
KP2_HEADER = ["Страна", "Город", "url", "Аккаунт", "Карта", "Статус",
              "Аккаунт", "Карта", "Статус"]


def kp2(*rows: list[str]) -> list[list[str]]:
    return [list(KP2_HEAD_TOP), list(KP2_HEADER), *[list(r) for r in rows]]


def test_status_columns() -> None:
    print("\n▸ Статусы: разбор блоков площадок")

    rows = kp2(
        ["Россия", "Москва", "https://metpromintex.ru",
         "https://yandex.ru/sprav/36559471/edit/posts/", "https://yandex.ru/maps/org/x/1/",
         "Активная",
         "account.2gis.com/orgs/70000001058778085/dashboard", "https://2gis.ru/moscow/", "Удалена"],
        ["Россия", "Казань", "https://kazan.metpromintex.ru",
         "https://yandex.ru/sprav/36347620/edit/posts/", "", "Онлайн", "", "", ""])
    got = A.parse_sheet(rows)
    cols = got["columns"]
    eq("ссылки Яндекса – колонка 3", cols["link"], 3)
    eq("ссылки 2ГИС – колонка 6", cols["gisLink"], 6)
    eq("статус Яндекса – свой", cols["status"], 5)
    eq("статус 2ГИС – свой", cols["gisStatus"], 8)
    eq("статус Яндекса прочитан", got["items"][0]["status"], "Активная")
    eq("статус 2ГИС прочитан", got["items"][0]["gisStatus"], "Удалена")
    eq("номер организации 2ГИС", A.gis_org_id(got["items"][0]["gisLink"]), "70000001058778085")
    eq("нет ссылки 2ГИС – нет номера", A.gis_org_id(got["items"][1]["gisLink"]), "")

    # Номер филиала с публичной карты кабинету не годится – его брать нельзя.
    eq("публичная карта – не кабинет",
       A.gis_org_id("https://2gis.ru/spb/inside/534866/firm/70000001081104279"), "")


def test_expected_state() -> None:
    print("\n▸ Статусы: чего ждём от площадки")

    eq("Активная – живая и опубликованная", A.expected_state("Активная"), "alive")
    eq("Онлайн – тоже живая", A.expected_state("Онлайн"), "alive")
    eq("Тех. проблемы – живая, публикация не важна",
       A.expected_state("Тех. проблемы"), "alive-any")
    eq("Удалена – карточки быть не должно", A.expected_state("Удалена"), "absent")
    eq("Добавить – карточки ещё нет", A.expected_state("Добавить"), "absent")
    eq("Заблокирована – карточки нет", A.expected_state("Заблокирована"), "absent")
    # «не активная» содержит «актив» – проверяться должна раньше, иначе
    # удалённая карточка сойдёт за живую.
    eq("не активная – карточки нет", A.expected_state("Не активная"), "absent")
    eq("пусто – сверять нечего", A.expected_state(""), "")
    eq("незнакомое – сверять нечего", A.expected_state("Что-то своё"), "")


def test_check_status() -> None:
    print("\n▸ Статусы: вердикты")

    alive = {"state": "alive", "published": True, "note": "живая"}
    unpub = {"state": "alive", "published": False, "note": "не опубликована"}
    gone = {"state": "absent", "published": None, "note": "карточки нет"}
    dunno = {"state": "unknown", "published": None, "note": "сессия слетела"}

    eq("Активная + живая = ок", A.check_status("Активная", alive)["verdict"], "ok")
    eq("Активная + нет карточки = ошибка", A.check_status("Активная", gone)["verdict"], "bad")
    eq("Активная + не опубликована = ошибка",
       A.check_status("Активная", unpub)["verdict"], "bad")
    eq("Удалена + нет карточки = ок", A.check_status("Удалена", gone)["verdict"], "ok")
    eq("Удалена + живая = ошибка", A.check_status("Удалена", alive)["verdict"], "bad")
    eq("Добавить + живая = ошибка", A.check_status("Добавить", alive)["verdict"], "bad")
    eq("Тех. проблемы + не опубликована = ок",
       A.check_status("Тех. проблемы", unpub)["verdict"], "ok")
    eq("Тех. проблемы + нет карточки = ошибка",
       A.check_status("Тех. проблемы", gone)["verdict"], "bad")

    # Не проверили – значит не проверили. Своя слепота не ошибка КП.
    eq("не проверено – не ошибка", A.check_status("Активная", dunno)["verdict"], "skip")
    eq("пустой статус – не сверяем", A.check_status("", alive)["verdict"], "skip")


def test_status_build() -> None:
    print("\n▸ Статусы: сборка отчёта")

    gis_link = "https://account.2gis.com/orgs/700001/dashboard"
    rows = kp2(
        # Всё сходится: карточка в списке, кабинет 2ГИС живой.
        ["Россия", "Шахты", "https://shahty.aviastal.ru",
         "https://yandex.ru/sprav/1/edit", "", "Активная", gis_link, "", "Активная"],
        # В КП «Удалена», а карточка живая – главное расхождение.
        ["Россия", "Казань", "https://kazan.aviastal.ru",
         "https://yandex.ru/sprav/2/edit", "", "Удалена", "", "", ""],
        # В КП «Активная», а Яндекс отдал 404 – второе расхождение.
        ["Россия", "Псков", "https://pskov.aviastal.ru",
         "https://yandex.ru/sprav/3/edit", "", "Активная", "", "", ""],
        # Статус есть, а проверить нечем: карточки нет ни в списке, ни в пробах.
        ["Россия", "Тверь", "https://tver.aviastal.ru",
         "https://yandex.ru/sprav/4/edit", "", "Активная", "", "", ""])
    cos = [company("1", "Ростовская область, Шахты", "https://shahty.aviastal.ru/"),
           company("2", "Республика Татарстан, Казань", "https://kazan.aviastal.ru/")]
    probes = {"3": {"state": "absent", "note": "карточка по ссылке не открылась (404)"}}
    states = {"700001": {"state": "alive", "note": "кабинет открывается"}}

    res = A.build(rows, cos, link_probes=probes, gis_states=states)
    t = res["totals"]
    # Шахты дают два совпадения – по Яндексу и по 2ГИС; у остальных городов
    # статус 2ГИС пуст, и в счёт идёт только Яндекс.
    eq("совпало две проверки", t["statusOk"], 2)
    eq("разошлось две", t["statusBad"], 2)
    eq("не проверили одну", t["statusUnchecked"], 1)

    bad = {(e["city"], e["platform"]) for e in A.status_entries(res, "bad")}
    eq("расхождения там, где ждали", bad, {("Казань", "Яндекс"), ("Псков", "Яндекс")})
    skip = [e["city"] for e in A.status_entries(res, "skip")]
    eq("не проверили Тверь", skip, ["Тверь"])
    # У Казани и Пскова 2ГИС-статус пуст – в «не проверено» они попасть не должны.
    check("пустой статус 2ГИС в список не лезет", "Казань" not in skip)

    sheet = A.status_rows(res)
    eq("в листе «Статусы» три строки", len(sheet) - 1, 3)
    eq("сначала расхождения", sheet[1][1], "Казань")
    check("непроверенное помечено отдельно", sheet[3][4].startswith("не проверено:"), sheet[3][4])

    # Ссылка на карточку рядом – иначе непонятно, что открывать.
    казань = next(e for e in A.status_entries(res, "bad") if e["city"] == "Казань")
    check("у расхождения есть ссылка", "/sprav/2/" in казань["url"], казань["url"])


def test_status_gis_states() -> None:
    print("\n▸ Статусы: 2ГИС по кабинетам")

    gis_link = "https://account.2gis.com/orgs/700002/dashboard"
    rows = kp2(["Россия", "Омск", "https://omsk.aviastal.ru", "", "", "", gis_link, "", "Активная"])

    def build(state):
        return A.build(rows, [], gis_states={"700002": state} if state else {})

    res = build({"state": "deleted", "note": "удалена из справочника"})
    eq("в КП активная, кабинет говорит «удалена» – ошибка", res["totals"]["statusBad"], 1)

    res = build({"state": "alive", "note": "кабинет открывается"})
    eq("кабинет живой – совпало", res["totals"]["statusOk"], 1)

    # Сессия слетела – это не ошибка КП, а наша слепота.
    res = build({"state": "login", "note": "страница входа"})
    eq("слетевшая сессия – не ошибка", res["totals"]["statusBad"], 0)
    eq("а «не проверено»", res["totals"]["statusUnchecked"], 1)
    note = A.status_entries(res, "skip")[0]["fact"]
    check("и понятно почему", "сесси" in note.lower(), note)

    res = build(None)
    eq("кабинет не обходили – не ошибка", res["totals"]["statusBad"], 0)
    eq("а «не проверено»", res["totals"]["statusUnchecked"], 1)


def test_status_real_kp() -> None:
    """
    Настоящая шапка КП МетПромИнтекс: три площадки, у всех колонка «Статус».

    Тут ломается всё, что ломается на живом файле: статус Google не должен
    сесть на 2ГИС, а статус 2ГИС – на Яндекс.
    """
    print("\n▸ Статусы: настоящая шапка КП")

    top = ["", "", "", "", "", "", "Телефония/Почта", "", "", "", "", "Мессенджеры", "",
           "Яндекс Бизнес", "", "", "", "", "2ГИС", "", "", "Google", "", ""]
    head = ["Страна", "Сортировка", "Город", "Численность", "url", "Адрес", "Почта",
            "Общий\nГород", "Общий\nСотовый", "Реклама\nГород", "SEO\nГород",
            "Telegram", "WhatsApp", "Посты", "Отгрузки", "Аккаунт", "Карта", "Статус",
            "Аккаунт", "Карта", "Статус", "Аккаунт", "Карта", "Статус"]
    row = ["Россия", "", "Екатеринбург", "1536000", "ekb.metpromintex.ru",
           "улица Кузнецова, 2Б", "ekaterinburg@metpromintex.ru", "+7 (343) 202-02-88",
           "", "", "", "", "", "", "",
           "https://yandex.ru/sprav/36902980/edit", "https://yandex.ru/maps/org/x/1/", "Активная",
           "https://account.2gis.com/orgs/70000001051648544/branches",
           "https://2gis.ru/ekaterinburg/search/x", "Удалена",
           "https://www.google.com/search?q=x", "https://www.google.com/maps/place/x", "Активная"]
    got = A.parse_sheet([top, head, row])
    it = got["items"][0]
    eq("статус Яндекса", it["status"], "Активная")
    eq("статус 2ГИС", it["gisStatus"], "Удалена")
    eq("кабинет 2ГИС", A.gis_org_id(it["gisLink"]), "70000001051648544")
    eq("город на месте", it["city"], "Екатеринбург")
    eq("сайт не перепутан со ссылкой площадки", it["site"], "ekb.metpromintex.ru")


def main() -> int:
    print("═" * 60)
    print("  ПРОВЕРКА СВЕРКИ КП С ЯНДЕКСОМ")
    print("═" * 60)
    test_normalize()
    test_parse_sheet()
    test_match()
    test_same_name_cities()
    test_duplicates()
    test_compare()
    test_export()
    test_report_sheets()
    test_yandex_pages()
    test_status_columns()
    test_expected_state()
    test_check_status()
    test_status_build()
    test_status_gis_states()
    test_status_real_kp()

    print("\n" + "═" * 60)
    if FAILED:
        print(f"  ПРОВАЛЕНО {len(FAILED)} из {PASSED + len(FAILED)}:")
        for name in FAILED:
            print(f"    • {name}")
        return 1
    print(f"  ВСЁ ХОРОШО – {PASSED} проверок пройдено")
    print("═" * 60)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
