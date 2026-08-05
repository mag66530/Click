"""
kp_sheet.py – города берутся из Google-таблицы «Карта присутствия» (КП).

Зачем: в облаке (Streamlit Cloud) файловая система временная – при перезапуске
контейнера список городов, набитый руками, пропадает. Таблица КП живёт снаружи,
поэтому источником правды делаем её: перезапустился контейнер – города
подтянулись обратно сами.

Доступ – через сервисный аккаунт Google (тот же, что в site-checker/kp_sheets.py):
таблица расшарена на его client_email как Читатель. Читаем ЗНАЧЕНИЯ ячеек через
Sheets API (values.get), а не скачиваем файл – так работает даже если у таблицы
выключено скачивание для читателей.

Что нужно положить в секреты приложения (Streamlit Cloud → Settings → Secrets):

    kp_sheet_url_SMU = "https://docs.google.com/spreadsheets/d/.../edit"
    kp_sheet_url_IMP = "..."
    kp_sheet_url_MPE = "..."

    # ключ сервисного аккаунта – одним из двух способов
    gcp_service_account_b64 = "<весь JSON-файл ключа в base64>"
    # ИЛИ секцией:
    # [gcp_service_account]
    # type = "service_account"
    # project_id = "..."
    # private_key = "-----BEGIN PRIVATE KEY-----\\n...\\n-----END PRIVATE KEY-----\\n"
    # client_email = "kp-reader@....iam.gserviceaccount.com"
    # ...

Ссылку на таблицу можно задать и в самом приложении («Города» → «Источник
городов»); тогда она сохранится в projects-config.json.

Как разбираем таблицу (структура КП, см. скриншот заказчика):
  • шапка – строка, где есть и «Страна», и «Город»;
  • колонка со ссылками Яндекс.Бизнеса определяется НЕ по названию (их в таблице
    несколько: «Аккаунт» есть и у Яндекса, и у 2ГИС, а шапка объединённая), а по
    самим значениям: берём колонку, где больше всего ссылок вида
    yandex.ru/sprav/<цифры>. Это устойчиво к перестановке колонок;
  • «Статус» – ближайшая колонка справа от неё; строки со статусом «Удалена»
    пропускаем.
"""

from __future__ import annotations

import base64
import json
import os
import re
from typing import Any

# Метка сборки: streamlit_app сверяет её и перезагружает модуль при расхождении.
BUILD = "2026-08-06-reviews-rules"

# Права нужны на оба API: у нативной таблицы значения читает Sheets API, а
# залитый .xlsx приходится качать файлом через Drive API.
SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly",
          "https://www.googleapis.com/auth/drive.readonly"]
GOOGLE_SHEET_MIME = "application/vnd.google-apps.spreadsheet"

SPRAV_RX = re.compile(r"yandex\.[a-z.]+/sprav/\d+", re.I)

# Значения колонки «Статус» у Яндекс.Бизнеса (список из выпадающего в КП):
# Активная, Онлайн, Тех. проблемы, Удалена, Добавить.
#   Удалена / Добавить  – карточки для работы нет, город не берём;
#   Тех. проблемы       – берём, но помечаем: молча пропустить город хуже,
#                         чем попробовать и увидеть ошибку в отчёте;
#   Активная / Онлайн / пусто – обычная работа.
# «не активная» – карточки для работы тоже нет. Проверяем ДО «актив», иначе
# «активные» и «не активные» неотличимы: в одной строке есть подстрока другой.
SKIP_STATUS_RX = re.compile(r"удал|добав|заблок|\bне\s*актив", re.I)
WARN_STATUS_RX = re.compile(r"тех|проблем", re.I)

# Одна и та же страна пишется в КП по-разному. Без сведения к одному имени
# Click рисует две плитки («РФ» и «Россия»), а окончания постов ищутся по
# названию страны – у «РФ» контактов не нашлось бы вовсе.
COUNTRY_ALIASES = {
    "рф": "Россия",
    "российская федерация": "Россия",
    "россия": "Россия",
    "казахстан": "Казахстан",
    "рк": "Казахстан",
    "беларусь": "Беларусь",
    "белоруссия": "Беларусь",
    "рб": "Беларусь",
}


def normalize_country(name: str) -> str:
    return COUNTRY_ALIASES.get(_norm(name), (name or "").strip())


# ─── Ссылка на таблицу ──────────────────────────────────────────────
def sheet_gid(url: str) -> str:
    """
    Номер листа из ссылки (`…/edit?gid=1223548196#gid=1223548196`).

    Заказчик присылает ссылку ИМЕННО на нужный лист – это самое точное
    указание, какое вообще бывает. Раньше мы его игнорировали и выбирали лист
    по названию: в таблице нашёлся старый «карта присутсвия Аэросталь (OLD» с
    четырьмя городами, и Click читал его вместо «кп», где городов 101.
    """
    if not url:
        return ""
    m = re.search(r"[?&#]gid=(\d+)", url)
    return m.group(1) if m else ""


def sheet_id(url: str) -> str:
    """ID таблицы из ссылки (или сам ID, если передали голым)."""
    if not url:
        return ""
    m = re.search(r"/spreadsheets/d/([A-Za-z0-9_\-]+)", url)
    if m:
        return m.group(1)
    url = url.strip()
    return url if re.fullmatch(r"[A-Za-z0-9_\-]{20,}", url or "") else ""


# Известные таблицы проектов. Ссылка сама по себе доступа не даёт – читать
# может только сервисный аккаунт, которому таблица расшарена. Держим их здесь,
# чтобы в облаке ничего не настраивать руками: файловая система там временная,
# и вписанная в приложении ссылка пропала бы при перезапуске.
DEFAULT_SHEET_URLS = {
    "MPI": "https://docs.google.com/spreadsheets/d/1Hb7R1scmyhSPyROUa2mWMzD4jDcdgBcs1IGu0WQ48O4/edit",
    "APS": ("https://docs.google.com/spreadsheets/d/1VCwHQRtxWzNv7yAK0pO4OHmxGOjrmI1f"
            "/edit?gid=266959061#gid=266959061"),
}


def sheet_url(project_id: str, from_config: str = "") -> str:
    """
    Приоритет: настройка в самом приложении → переменная окружения →
    секрет kp_sheet_url_<ПРОЕКТ> → общий секрет kp_sheet_url → таблица проекта
    по умолчанию.
    """
    if (from_config or "").strip():
        return from_config.strip()
    env = (os.environ.get(f"kp_sheet_url_{project_id}") or os.environ.get("kp_sheet_url") or "").strip()
    if env:
        return env
    try:
        import streamlit as st
        for key in (f"kp_sheet_url_{project_id}", "kp_sheet_url"):
            v = st.secrets.get(key)
            if v:
                return str(v).strip()
    except Exception:
        pass
    return DEFAULT_SHEET_URLS.get(project_id, "")


# ─── Ключ сервисного аккаунта ───────────────────────────────────────
def _from_b64(value: str) -> dict:
    return json.loads(base64.b64decode(value.strip()).decode("utf-8"))


def service_account_info() -> dict | None:
    """
    JSON-ключ сервисного аккаунта. Источники по порядку:
      1) окружение GCP_SA_JSON (строка-JSON) или GCP_SA_B64 (base64);
      2) секреты Streamlit: gcp_service_account (секция или строка-JSON)
         либо gcp_service_account_b64 (base64 всего файла – удобнее для TOML).
    None – ключ не задан или не разобрался.
    """
    raw = (os.environ.get("GCP_SA_JSON") or "").strip()
    if raw:
        try:
            return json.loads(raw)
        except Exception:
            return None
    b64env = (os.environ.get("GCP_SA_B64") or "").strip()
    if b64env:
        try:
            return _from_b64(b64env)
        except Exception:
            return None
    try:
        import streamlit as st
    except Exception:
        return None
    try:
        v = st.secrets.get("gcp_service_account")
    except Exception:
        v = None
    if v is not None:
        try:
            return json.loads(v) if isinstance(v, str) else dict(v)
        except Exception:
            pass
    try:
        b = st.secrets.get("gcp_service_account_b64")
    except Exception:
        b = None
    if b:
        try:
            return _from_b64(str(b))
        except Exception:
            pass
    return None


def is_configured(project_id: str, from_config: str = "") -> bool:
    return bool(sheet_id(sheet_url(project_id, from_config))) and service_account_info() is not None


# ─── Чтение таблицы ─────────────────────────────────────────────────
def _auth_headers(sa_info: dict) -> dict:
    from google.auth.transport.requests import Request as GARequest
    from google.oauth2 import service_account

    creds = service_account.Credentials.from_service_account_info(sa_info, scopes=SCOPES)
    creds.refresh(GARequest())
    return {"Authorization": f"Bearer {creds.token}"}


def _file_mime(file_id: str, headers: dict) -> str:
    """Что это за файл на Диске: нативная таблица или залитый Excel."""
    import requests
    r = requests.get(f"https://www.googleapis.com/drive/v3/files/{file_id}",
                     params={"fields": "mimeType,name", "supportsAllDrives": "true"},
                     headers=headers, timeout=30)
    if r.status_code != 200:
        return ""          # не смогли определить – попробуем как обычную таблицу
    return (r.json() or {}).get("mimeType", "")


def _read_uploaded_xlsx(file_id: str, headers: dict) -> dict[str, list[list[str]]]:
    """
    Листы Excel-файла, ЗАЛИТОГО на Google Диск.

    Такой файл открывается в Google Таблицах и по ссылке неотличим от обычной
    таблицы, но Sheets API его не читает («must not be an Office file»).
    Скачиваем сам файл через Drive API и разбираем – ровно как это давно
    делает site-checker, поэтому у него такие КП и читались.
    """
    import io

    import requests
    from openpyxl import load_workbook

    r = requests.get(f"https://www.googleapis.com/drive/v3/files/{file_id}",
                     params={"alt": "media", "supportsAllDrives": "true"},
                     headers=headers, timeout=90)
    if r.status_code != 200:
        raise RuntimeError(f"Drive API вернул HTTP {r.status_code}: {r.text[:160]} "
                           "(расшарена ли таблица на сервисный аккаунт? включён ли Drive API?)")
    wb = load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
    out: dict[str, list[list[str]]] = {}
    for ws in wb.worksheets:
        out[ws.title] = [[("" if c.value is None else str(c.value)) for c in row]
                         for row in ws.iter_rows()]
    wb.close()
    return out


def _read_values(file_id: str, sa_info: dict, gid: str = "") -> list[list[str]]:
    """Значения таблицы. Нативную читаем через Sheets API, залитый .xlsx – качаем."""
    import requests

    headers = _auth_headers(sa_info)
    base = f"https://sheets.googleapis.com/v4/spreadsheets/{file_id}"

    # Залитый Excel Sheets API не осилит – сразу идём в Drive.
    if _file_mime(file_id, headers) not in (GOOGLE_SHEET_MIME, ""):
        sheets = _read_uploaded_xlsx(file_id, headers)
        return _pick_sheet(list(sheets), lambda t: sheets[t])

    meta = requests.get(base, params={"fields": "sheets.properties(sheetId,title)"},
                        headers=headers, timeout=30)
    if meta.status_code == 403:
        raise RuntimeError(
            f"Google отказал (403). Расшарена ли таблица на {sa_info.get('client_email', 'сервисный аккаунт')} "
            "как Читатель и включён ли Google Sheets API в проекте?"
        )
    if meta.status_code != 200:
        # Тип файла определить не удалось, а Sheets API упёрся в Office-файл –
        # значит это всё-таки залитый Excel, качаем его напрямую.
        if "must not be an Office file" in (meta.text or ""):
            sheets = _read_uploaded_xlsx(file_id, headers)
            return _pick_sheet(list(sheets), lambda t: sheets[t])
        raise RuntimeError(f"Sheets API вернул HTTP {meta.status_code}: {meta.text[:160]}")

    props = [sh.get("properties", {}) for sh in (meta.json() or {}).get("sheets", [])]
    titles = [p.get("title", "") for p in props if p.get("title")]
    if not titles:
        raise RuntimeError("В таблице нет листов")
    # Ссылка с gid указывает на конкретный лист – начинаем с него.
    prefer = next((p.get("title", "") for p in props
                   if gid and str(p.get("sheetId")) == str(gid)), "")

    def read(title: str) -> list[list[str]]:
        rng = requests.utils.quote(f"'{title}'")
        r = requests.get(f"{base}/values/{rng}",
                         params={"majorDimension": "ROWS", "valueRenderOption": "FORMATTED_VALUE"},
                         headers=headers, timeout=60)
        if r.status_code != 200:
            raise RuntimeError(f"Не удалось прочитать лист «{title}»: HTTP {r.status_code}")
        return [[("" if v is None else str(v)) for v in row] for row in (r.json() or {}).get("values", [])]

    return _pick_sheet(titles, read, prefer)


def _pick_sheet(titles: list[str], read, prefer: str = "") -> list[list[str]]:
    """Первый лист, на котором нашлись города со ссылками на Яндекс.Бизнес."""
    if not titles:
        raise RuntimeError("В таблице нет листов")
    # Старые копии не читаем ВООБЩЕ – ровно этого просила заказчик. К ним
    # возвращаемся, только если на нормальных листах городов не нашлось: иначе
    # таблица, где рабочий лист один и он назван «…OLD», перестала бы читаться.
    normal = [t for t in titles if t == prefer or not _OLD_SHEET_RX.search((t or "").lower())]
    skipped_old = [t for t in titles if t not in normal]

    tried = []
    for title in _sheet_order(normal, prefer) + _sheet_order(skipped_old, prefer):
        try:
            rows = read(title)
        except Exception as e:  # noqa: BLE001
            tried.append(f"«{title}»: {e}")
            continue
        cities, diag = parse_rows(rows)
        if cities:
            return rows
        tried.append(f"«{title}»: {diag.get('error') or 'городов не нашлось'}")
    raise RuntimeError("Ни на одном листе не нашлось таблицы с городами и ссылками "
                       "на Яндекс.Бизнес. Проверено – " + "; ".join(tried[:6]))


# Листы-двойники: старые копии, оригиналы «не трогать», архивы. Названия у них
# похожи на рабочие, поэтому по имени их легко перепутать – отправляем в конец.
_OLD_SHEET_RX = re.compile(r"\bold\b|ориг|архив|копия|старая|старый|backup|не\s*трогать", re.I)


def _sheet_order(titles: list[str], prefer: str = "") -> list[str]:
    """
    Порядок обхода листов. Первым идёт лист, на который прямо указали ссылкой
    (gid), затем «КП» и «Карта присутствия» – в таблице города именно там.
    Название сверяем без пробелов и с допуском на опечатку: в реальном файле
    лист называется «Карта присутсвия».
    """
    def key(t: str) -> tuple[int, int]:
        raw = (t or "").lower()
        n = re.sub(r"[^а-яё]", "", raw)
        old = 10 if _OLD_SHEET_RX.search(raw) else 0
        if prefer and t == prefer:
            return (-1, 0)                 # указан ссылкой – важнее всего
        if n in {"кп", "кпшка"}:
            return (old + 0, 1)
        if n.startswith("картаприсут"):
            return (old + 0, 2)
        if "присут" in n or "город" in n:
            return (old + 1, 0)
        if n in {"сводка", "нетрогать", "приоритеты"}:
            return (old + 3, 0)
        return (old + 2, 0)

    return sorted(titles, key=key)


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip().lower()


def parse_rows(rows: list[list[str]]) -> tuple[list[dict], dict]:
    """
    Разбирает значения таблицы в список городов.
    Возвращает (города, диагностика). Города: [{country, name, url, status}].

    Вынесено отдельно от сети, чтобы можно было проверить тестом.
    """
    diag: dict[str, Any] = {"rows": len(rows), "skippedDeleted": 0, "skippedNoUrl": 0}
    if not rows:
        return [], diag

    # 1. Шапка – строка, где есть и «Страна», и «Город»
    header_idx = -1
    for i, row in enumerate(rows[:20]):
        cells = [_norm(c) for c in row]
        if "страна" in cells and "город" in cells:
            header_idx = i
            break
    if header_idx < 0:
        diag["error"] = "Не нашли шапку: нужна строка с колонками «Страна» и «Город»"
        return [], diag

    header = [_norm(c) for c in rows[header_idx]]
    col_country = header.index("страна")
    col_city = header.index("город")
    data = rows[header_idx + 1:]

    # 2. Колонка со ссылками ЯБ – по содержимому, а не по названию:
    #    «Аккаунт» в таблице есть и у Яндекса, и у 2ГИС.
    width = max((len(r) for r in data), default=0)
    best_col, best_hits = -1, 0
    for c in range(width):
        hits = sum(1 for r in data if c < len(r) and SPRAV_RX.search(r[c]))
        if hits > best_hits:
            best_col, best_hits = c, hits
    if best_col < 0:
        diag["error"] = "В таблице не нашли колонку со ссылками вида yandex.ru/sprav/…"
        return [], diag
    diag["urlColumn"] = best_col
    diag["urlHeader"] = rows[header_idx][best_col] if best_col < len(rows[header_idx]) else ""

    # 2б. Контакты города – слева от блоков площадок, названия однозначные.
    def find(*names: str) -> int:
        for want in names:
            for c, h in enumerate(header):
                if h == want:
                    return c
        return -1

    col_site = find("url", "сайт")
    col_email = find("почта", "email", "e-mail")
    col_phone = find("общий город", "общий\nгород", "телефон")
    diag["contactColumns"] = {"site": col_site, "email": col_email, "phone": col_phone}

    # 3. «Статус» – ближайшая колонка справа от ссылок (у Яндекса свой, у 2ГИС свой).
    #    Название бывает и «Статус», и «Яндекс_Статус» – ищем по вхождению, но
    #    отсекаем чужие площадки, чтобы не подхватить статус 2ГИС или Гугла.
    col_status = -1
    for c in range(best_col + 1, min(best_col + 6, len(header))):
        h = header[c]
        if "статус" in h and not re.search(r"2гис|2gis|гугл|google", h):
            col_status = c
            break
    diag["statusColumn"] = col_status
    diag["statusHeader"] = rows[header_idx][col_status] if 0 <= col_status < len(rows[header_idx]) else ""

    cities: list[dict] = []
    for r in data:
        def cell(i: int) -> str:
            return r[i].strip() if 0 <= i < len(r) else ""

        url_raw = cell(best_col)
        m = SPRAV_RX.search(url_raw)
        if not m:
            continue
        country, name = cell(col_country), cell(col_city)
        if not country or not name:
            diag["skippedNoUrl"] += 1
            continue
        status = cell(col_status) if col_status >= 0 else ""
        if status and SKIP_STATUS_RX.search(status):
            diag["skippedDeleted"] += 1
            continue
        if status and WARN_STATUS_RX.search(status):
            diag["withProblems"] = diag.get("withProblems", 0) + 1
        cities.append({
            "country": normalize_country(country),
            "name": name,
            "url": "https://" + m.group(0) if not url_raw.lower().startswith("http") else url_raw,
            "status": status,
            # Контакты из той же строки КП – ими можно подставлять окончания постов,
            # чтобы номер и почта менялись правкой таблицы, а не кода.
            "site": cell(col_site),
            "email": cell(col_email),
            "phone": cell(col_phone),
        })

    diag["cities"] = len(cities)
    diag["countries"] = len({c["country"] for c in cities})
    return cities, diag


def load_cities(project_id: str, from_config: str = "") -> tuple[list[dict], dict]:
    """Скачать и разобрать КП проекта. Бросает RuntimeError с понятным текстом."""
    url = sheet_url(project_id, from_config)
    fid = sheet_id(url)
    if not fid:
        raise RuntimeError("Не задана ссылка на Google-таблицу КП для этого проекта.")
    sa = service_account_info()
    if sa is None:
        raise RuntimeError(
            "Не найден ключ сервисного аккаунта Google. Добавьте в секреты приложения "
            "gcp_service_account_b64 (весь JSON-ключ в base64) или секцию [gcp_service_account]."
        )
    return parse_rows(_read_values(fid, sa, sheet_gid(url)))


def to_countries(cities: list[dict], project_id: str) -> list[dict]:
    """
    Список городов → структура конфига Click: [{id, name, cities:[{id,name,url}]}].
    Порядок стран и городов сохраняется как в таблице.
    """
    def slug(s: str) -> str:
        return re.sub(r"[^0-9a-zа-яё]+", "-", (s or "").lower()).strip("-") or "x"

    by_country: dict[str, list[dict]] = {}
    for c in cities:
        by_country.setdefault(c["country"], []).append(c)

    out = []
    for n, (country, items) in enumerate(by_country.items()):
        out.append({
            "id": f"c-{project_id}-{n}-{slug(country)}",
            "name": country,
            "cities": [
                {"id": f"ct-{project_id}-{n}-{i}-{slug(it['name'])}", "name": it["name"], "url": it["url"]}
                for i, it in enumerate(items)
            ],
        })
    return out
